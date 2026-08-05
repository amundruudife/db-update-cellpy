import pytest
from openpyxl import load_workbook

from src.contracts import LOCAL_SNAPSHOT_PATH, SOURCE_SHEET_NAME
from src import snapshot as snapshot_module
from src.snapshot import snapshot_path, write_snapshot
from src.source_validation import SourceValidationError


def _source_rows():
    return [
        ["metadata", *([None] * 18)],
        ["key", *([None] * 18)],
        ["units", *([None] * 18)],
        [
            101,
            "channel",
            "cell",
            "file.xlsx",
            1.0,
            2.0,
            3.0,
            4.0,
            5.0,
            "constant",
            6.0,
            "schedule",
            7.0,
            8.0,
            9.0,
            10.0,
            11.0,
            12.0,
            "#N/A",
        ],
    ]


def test_snapshot_contains_only_the_hard_coded_cp_sheet(tmp_path):
    snapshot_path = write_snapshot(_source_rows(), root=tmp_path)

    workbook = load_workbook(snapshot_path, data_only=False, read_only=True)
    try:
        assert workbook.sheetnames == [SOURCE_SHEET_NAME]
        worksheet = workbook[SOURCE_SHEET_NAME]
        assert worksheet.max_row == 4
        assert worksheet.max_column == 19
        assert worksheet["A1"].value == "metadata"
        assert worksheet["A2"].value == "key"
        assert worksheet["A4"].value == 101
        assert worksheet["S4"].value == "#N/A"
    finally:
        workbook.close()


def test_snapshot_path_is_the_fixed_source_data_cell_log_cp_path(tmp_path):
    assert snapshot_path(tmp_path) == tmp_path / LOCAL_SNAPSHOT_PATH
    assert str(LOCAL_SNAPSHOT_PATH).replace("\\", "/") == "source_data/Cell_Log_CP.xlsx"


def test_failed_validation_preserves_the_last_accepted_snapshot(tmp_path):
    destination = write_snapshot(_source_rows(), root=tmp_path)
    before = destination.read_bytes()
    invalid = _source_rows()
    invalid.append(invalid[-1].copy())

    with pytest.raises(SourceValidationError, match="duplicate_ids"):
        write_snapshot(invalid, root=tmp_path)

    assert destination.read_bytes() == before


def test_replace_failure_preserves_snapshot_and_cleans_temporary_file(tmp_path, monkeypatch):
    destination = write_snapshot(_source_rows(), root=tmp_path)
    before = destination.read_bytes()
    replace_calls = []
    original_replace = snapshot_module.os.replace

    def fail_replace(source, target):
        replace_calls.append((source, target))
        raise OSError("simulated replace failure")

    monkeypatch.setattr(snapshot_module.os, "replace", fail_replace)
    with pytest.raises(OSError, match="simulated replace failure"):
        write_snapshot(_source_rows(), root=tmp_path)

    assert destination.read_bytes() == before
    assert len(replace_calls) == 1
    temporary, target = replace_calls[0]
    assert target == destination
    assert temporary.parent == destination.parent
    assert not temporary.exists()

    monkeypatch.setattr(snapshot_module.os, "replace", original_replace)
    reopened = load_workbook(destination, data_only=True, read_only=True)
    try:
        assert reopened[SOURCE_SHEET_NAME]["A4"].value == 101
    finally:
        reopened.close()
