from pathlib import Path

from openpyxl import Workbook

from src.database_inventory import (
    classify_cell,
    compare_cell_inventories,
    inventory_workbook,
)


def make_fixture(path: Path) -> None:
    workbook = Workbook()
    mirror = workbook.active
    mirror.title = "Slurry"
    mirror["A1"] = "id"
    mirror["A3"] = 100
    mirror["M3"] = "preserve"
    database = workbook.create_sheet("db_table")
    database["A1"] = "id"
    database["A3"] = "=Slurry!A3"
    database["F3"] = "manual"
    database["M3"] = "legacy"
    database.merge_cells("B1:C1")
    database.auto_filter.ref = "A1:M2"
    database.freeze_panes = "A2"
    database.column_dimensions["A"].width = 14
    database.row_dimensions[2].height = 22
    workbook.save(path)


def test_inventory_captures_non_formula_cells_outside_manual_columns_and_topology(tmp_path):
    path = tmp_path / "fixture.xlsx"
    make_fixture(path)

    report = inventory_workbook(path)

    assert report["sheet_order"] == ["Slurry", "db_table"]
    assert report["sheets"]["db_table"]["dimensions"] == "A1:M3"
    assert report["sheets"]["db_table"]["merges"] == ["B1:C1"]
    assert report["sheets"]["db_table"]["auto_filter"] == "A1:M2"
    assert report["sheets"]["db_table"]["freeze_panes"] == "A2"
    assert {cell["coordinate"] for cell in report["cells"]} == {"M3"}
    assert report["cells"][0]["classification"] == "retain_existing"
    assert report["column_summary"]["M"]["non_formula_populated"] == 1


def test_cell_classification_is_explicit_and_comparison_reports_deletions():
    assert classify_cell("=A1", is_formula=True) == "recreate_formula"
    assert classify_cell("legacy", is_formula=False) == "retain_existing"
    assert classify_cell(None, is_formula=False) == "delete"

    before = [{"sheet": "db_table", "coordinate": "M3", "value": "legacy"}]
    after = []
    changes = compare_cell_inventories(before, after)
    assert changes == [{"sheet": "db_table", "coordinate": "M3", "classification": "delete", "before": "legacy"}]


def test_inventory_records_row_offset_evidence(tmp_path):
    path = tmp_path / "fixture.xlsx"
    make_fixture(path)

    report = inventory_workbook(path, row_offset=("db_table", "Slurry", 3, 2))

    assert report["row_offset"] == {
        "target_sheet": "db_table",
        "source_sheet": "Slurry",
        "target_row": 3,
        "source_row": 2,
        "offset": -1,
    }
