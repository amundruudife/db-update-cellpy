import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.candidate_pipeline import CandidateBuildError, build_candidate
import src.candidate_pipeline as candidate_pipeline
from src.neware import read_neware_source


NEWARE_HEADERS = [
    "Test No",
    "Cell ID",
    "Cell batch",
    "Project",
    "Cell type",
    "Cell test label",
    "Cell capacity (Ah)",
    "Test Schedule",
    "Test temp (°C)",
    "Comment",
    "Ignored column",
]


def _write_neware(path: Path, rows, *, with_other_sheet=True):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "test_log"
    for column, header in enumerate(NEWARE_HEADERS, start=1):
        sheet.cell(1, column, header)
        sheet.cell(2, column, "str")
    for row_number, row in enumerate(rows, start=3):
        for column, header in enumerate(NEWARE_HEADERS, start=1):
            sheet.cell(row_number, column, row.get(header))
    if with_other_sheet:
        other = workbook.create_sheet("other")
        other.append(NEWARE_HEADERS)
        other.cell(2, NEWARE_HEADERS.index("Cell test label") + 1, "must-not-be-read")
    workbook.save(path)
    workbook.close()


def _write_legacy_source(path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "log"
    sheet["A1"] = "key"
    sheet["C1"] = "proj"
    sheet["A5"] = 101
    sheet["C5"] = "CellMap"
    workbook.save(path)
    workbook.close()


def _write_database(path: Path, *, ids=(101,)):
    workbook = Workbook()
    slurry = workbook.active
    slurry.title = "Slurry"
    for column in range(1, 88):
        slurry.cell(1, column, f"header-{column}")
        slurry.cell(2, column, f"unit-{column}")
        slurry.cell(3, column, f"field-{column}")
    database = workbook.create_sheet("db_table")
    for column in range(1, 74):
        database.cell(1, column, f"db-header-{column}")
        database.cell(2, column, f"db-type-{column}")
    database["A1"] = "id"
    database["D1"] = "exists"
    for offset, record_id in enumerate(ids, start=4):
        slurry.cell(offset, 1, record_id)
        database.cell(offset - 1, 1, record_id)
    workbook.save(path)
    workbook.close()


def _row(test_no, label, *, cell_id="CELL-1", batch="batch-1", project="Project"):
    return {
        "Test No": test_no,
        "Cell ID": cell_id,
        "Cell batch": batch,
        "Project": project,
        "Cell type": "pouch",
        "Cell test label": label,
        "Cell capacity (Ah)": 1.25,
        "Test Schedule": "C/10",
        "Test temp (°C)": 25,
        "Comment": None,
        "Ignored column": "ignore-me",
    }


def test_neware_reader_uses_test_log_only_and_reports_exact_placeholders(tmp_path):
    source = tmp_path / "Neware_log.xlsx"
    placeholder = _row(1, "template-one", cell_id=" Cell ID or Cell serial number")
    placeholder_two = _row(2, "template-two", cell_id=" Cell ID or Cell serial number")
    usable = _row(34, "label-1")
    usable_two = _row(34, "label-2")
    _write_neware(source, [placeholder, placeholder_two, usable, usable_two])

    result = read_neware_source(source)

    assert result.source_rows == 4
    assert result.placeholder_rows == (3, 4)
    assert [row.cell_test_label for row in result.rows] == ["label-1", "label-2"]
    assert result.rows[0].test_no == result.rows[1].test_no == 34
    assert result.rows[0].mapped_values() == {
        "label": "CELL-1",
        "batch": "batch-1",
        "project": "Project",
        "cell_type": "pouch",
        "cell": "label-1",
        "nominal_capacity": 1.25,
        "schedule": "C/10",
        "temperature": 25,
        "comment_general": None,
    }


def test_neware_rows_append_directly_and_manifest_is_reused(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    neware = tmp_path / "neware.xlsx"
    manifest = tmp_path / "neware_id_manifest.json"
    candidate = tmp_path / "candidate.xlsx"
    rerun = tmp_path / "rerun.xlsx"
    _write_legacy_source(source)
    _write_database(database)
    _write_neware(neware, [_row(34, "label-1"), _row(35, "label-2", cell_id="CELL-2")])

    report = build_candidate(
        source,
        database,
        candidate,
        neware_source_path=neware,
        neware_manifest_path=manifest,
    )

    assert report.neware_new_ids == (102, 103)
    assert report.neware_placeholder_rows == ()
    workbook = load_workbook(candidate, read_only=True, data_only=False)
    try:
        database_sheet = workbook["db_table"]
        assert database_sheet.max_row == 5
        assert [database_sheet.cell(4, column).value for column in (1, 3, 4, 18, 20, 22, 23, 24, 26, 57, 59, 62)] == [
            102,
            "batch-1",
            1,
            "pouch",
            "arbin_sql_h5",
            "Project",
            "CELL-1",
            "cycling",
            "label-1",
            "C/10",
            None,
            25,
        ]
        assert database_sheet["AA4"].value is None
    finally:
        workbook.close()

    manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
    assert manifest_data["entries"]["label-1"]["id"] == 102

    rerun_report = build_candidate(
        source,
        candidate,
        rerun,
        neware_source_path=neware,
        neware_manifest_path=manifest,
    )
    assert rerun_report.neware_new_ids == ()
    assert rerun_report.neware_retained_ids == (102, 103)
    rerun_workbook = load_workbook(rerun, read_only=True)
    try:
        assert rerun_workbook["db_table"].max_row == 5
    finally:
        rerun_workbook.close()


@pytest.mark.parametrize("failure_at", [1, 2, 3], ids=["before-manifest", "before-candidate", "before-report"])
def test_publication_interruption_is_ordered_restart_recoverable_and_idempotent(
    tmp_path, monkeypatch, failure_at
):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    neware = tmp_path / "neware.xlsx"
    manifest = tmp_path / "manifest.json"
    candidate = tmp_path / "candidate.xlsx"
    report = tmp_path / "report.json"
    _write_legacy_source(source)
    _write_database(database)
    _write_neware(neware, [_row(1, "label-1")])
    build_candidate(
        source,
        database,
        candidate,
        report_path=report,
        neware_source_path=neware,
        neware_manifest_path=manifest,
    )
    old = {path: path.read_bytes() for path in (manifest, candidate, report)}
    _write_neware(neware, [_row(1, "label-1"), _row(2, "label-2", cell_id="CELL-2")])

    calls = []
    original_publish = candidate_pipeline._publish_outputs

    def interrupted_publish(*args, **kwargs):
        original_replace = candidate_pipeline.os.replace

        def interrupted_replace(source_path, destination_path):
            calls.append(Path(destination_path).name)
            if len(calls) == failure_at:
                raise OSError("injected publication interruption")
            return original_replace(source_path, destination_path)

        monkeypatch.setattr(candidate_pipeline.os, "replace", interrupted_replace)
        try:
            return original_publish(*args, **kwargs)
        finally:
            monkeypatch.setattr(candidate_pipeline.os, "replace", original_replace)

    monkeypatch.setattr(candidate_pipeline, "_publish_outputs", interrupted_publish)
    with pytest.raises(OSError, match="interruption"):
        build_candidate(
            source,
            database,
            candidate,
            report_path=report,
            neware_source_path=neware,
            neware_manifest_path=manifest,
        )
    monkeypatch.setattr(candidate_pipeline, "_publish_outputs", original_publish)

    assert calls == ["manifest.json", "candidate.xlsx", "report.json"][:failure_at]
    assert all(path.exists() for path in (manifest, candidate, report))
    for index, path in enumerate((manifest, candidate, report), start=1):
        if index >= failure_at:
            assert path.read_bytes() == old[path]
        else:
            assert path.read_bytes() != old[path]
    assert not any(
        path.name.startswith((".candidate", ".report", ".manifest"))
        for path in tmp_path.iterdir()
    )

    rerun = build_candidate(
        source,
        database,
        candidate,
        report_path=report,
        neware_source_path=neware,
        neware_manifest_path=manifest,
    )
    assert rerun.neware_new_ids == (102, 103)
    assert json.loads(manifest.read_text(encoding="utf-8"))["entries"]["label-2"]["id"] == 103
    assert not any(
        path.name.startswith((".candidate", ".report", ".manifest"))
        for path in tmp_path.iterdir()
    )


def test_neware_changed_payload_and_duplicate_label_fail_closed(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    neware = tmp_path / "neware.xlsx"
    manifest = tmp_path / "neware_id_manifest.json"
    candidate = tmp_path / "candidate.xlsx"
    _write_legacy_source(source)
    _write_database(database)
    _write_neware(neware, [_row(1, "label-1"), _row(2, "label-1")])

    with pytest.raises(CandidateBuildError, match="duplicate Cell test label"):
        build_candidate(
            source,
            database,
            candidate,
            neware_source_path=neware,
            neware_manifest_path=manifest,
        )
    assert not manifest.exists()


def test_neware_existing_payload_change_and_cellpy_boundary_fail_closed(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    neware = tmp_path / "neware.xlsx"
    manifest = tmp_path / "neware_id_manifest.json"
    candidate = tmp_path / "candidate.xlsx"
    changed_candidate = tmp_path / "changed.xlsx"
    _write_legacy_source(source)
    _write_database(database)
    _write_neware(neware, [_row(1, "label-1")])

    build_candidate(
        source,
        database,
        candidate,
        neware_source_path=neware,
        neware_manifest_path=manifest,
    )
    manifest_before = manifest.read_bytes()

    workbook = load_workbook(neware)
    try:
        workbook["test_log"]["G3"] = 9.9
        workbook.save(neware)
    finally:
        workbook.close()

    with pytest.raises(CandidateBuildError, match="changed existing Neware payload"):
        build_candidate(
            source,
            candidate,
            changed_candidate,
            neware_source_path=neware,
            neware_manifest_path=manifest,
        )
    assert not changed_candidate.exists()
    assert manifest.read_bytes() == manifest_before

    with pytest.raises(CandidateBuildError, match="cannot be combined"):
        build_candidate(
            source,
            database,
            tmp_path / "cellpy.xlsx",
            neware_source_path=neware,
            neware_manifest_path=tmp_path / "cellpy-manifest.json",
            cellpy_ready=True,
        )


def test_neware_manifest_id_collision_with_slurry_fails_closed(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    neware = tmp_path / "neware.xlsx"
    manifest = tmp_path / "neware_id_manifest.json"
    candidate = tmp_path / "candidate.xlsx"
    _write_legacy_source(source)
    _write_database(database)
    _write_neware(neware, [_row(1, "label-1")])

    build_candidate(
        source,
        database,
        candidate,
        neware_source_path=neware,
        neware_manifest_path=manifest,
    )
    document = json.loads(manifest.read_text(encoding="utf-8"))
    document["entries"]["label-1"]["id"] = 101
    manifest.write_text(json.dumps(document), encoding="utf-8")

    with pytest.raises(CandidateBuildError, match="collides with an existing Slurry ID"):
        build_candidate(
            source,
            database,
            tmp_path / "collision.xlsx",
            neware_source_path=neware,
            neware_manifest_path=manifest,
        )


def test_neware_reservation_reconciles_unreferenced_slurry_id_before_allocation(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    neware = tmp_path / "neware.xlsx"
    manifest = tmp_path / "neware_id_manifest.json"
    candidate = tmp_path / "candidate.xlsx"
    _write_legacy_source(source)
    _write_database(database)
    workbook = load_workbook(database)
    try:
        workbook["Slurry"]["A5"] = 102
        workbook.save(database)
    finally:
        workbook.close()
    _write_neware(neware, [_row(1, "label-1")])

    with pytest.raises(CandidateBuildError, match="missing Slurry IDs: 102"):
        build_candidate(
            source,
            database,
            candidate,
            neware_source_path=neware,
            neware_manifest_path=manifest,
        )
    assert not manifest.exists()


def test_manifest_ahead_requires_matching_neware_source_row(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    neware = tmp_path / "neware.xlsx"
    manifest = tmp_path / "neware_id_manifest.json"
    candidate = tmp_path / "candidate.xlsx"
    _write_legacy_source(source)
    _write_database(database)
    _write_neware(neware, [_row(1, "label-1")])
    build_candidate(
        source,
        database,
        candidate,
        neware_source_path=neware,
        neware_manifest_path=manifest,
    )
    document = json.loads(manifest.read_text(encoding="utf-8"))
    payload = _row(2, "label-2", cell_id="CELL-2")
    payload.pop("Ignored column")
    document["entries"]["label-2"] = {"id": 103, "payload": payload}
    manifest.write_text(json.dumps(document), encoding="utf-8")
    _write_neware(neware, [_row(1, "label-1")])

    with pytest.raises(CandidateBuildError, match="manifest-ahead"):
        build_candidate(
            source,
            database,
            tmp_path / "ahead.xlsx",
            neware_source_path=neware,
            neware_manifest_path=manifest,
        )
