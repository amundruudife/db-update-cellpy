import builtins
import json
import os
import tempfile
import zipfile
from pathlib import Path
from types import SimpleNamespace
from xml.etree import ElementTree

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from openpyxl.workbook.defined_name import DefinedName

import src.candidate_pipeline as candidate_pipeline
from src.contracts import PRODUCTION_DATABASE_PATH
from src.candidate_pipeline import CandidateBuildError, build_candidate


PROJECT = "CellMap"


def _source(path: Path, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "log"
    for column in range(1, 88):
        sheet.cell(1, column, f"group-{column}")
        sheet.cell(2, column, f"header-{column}")
        sheet.cell(3, column, f"unit-{column}")
        sheet.cell(4, column, f"field-{column}")
    sheet["A2"] = "key"
    sheet["C2"] = "proj"
    sheet["A4"] = "key"
    sheet["C4"] = "proj"
    for row_number, (record_id, project, marker) in enumerate(rows, start=5):
        sheet.cell(row_number, 1, record_id)
        sheet.cell(row_number, 3, project)
        sheet.cell(row_number, 4, marker)
    workbook.save(path)
    workbook.close()


def _database(path: Path, ids=(101,)):
    workbook = Workbook()
    slurry = workbook.active
    slurry.title = "Slurry"
    for column in range(1, 88):
        slurry.cell(1, column, f"header-{column}")
        slurry.cell(2, column, f"unit-{column}")
        slurry.cell(3, column, f"field-{column}")

    database = workbook.create_sheet("db_table")
    for column in range(1, 74):
        database.cell(1, column, f"db-header-{column}")
        database.cell(2, column, f"db-type-{column}")
    database["A1"] = "id"
    database["D1"] = "exists"
    for offset, record_id in enumerate(ids, start=4):
        slurry.cell(offset, 1, record_id)
        slurry.cell(offset, 4, f"existing-{record_id}")
        db_row = offset - 1
        database.cell(db_row, 1, record_id)
        for column in range(6, 13):
            database.cell(db_row, column, f"manual-{record_id}-{column}")
        database.cell(db_row, 30, f"historical-{record_id}")
    workbook.save(path)
    workbook.close()


def _add_data_validation_extension(path: Path):
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    x14_ns = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
    extension_uri = "{CCE6A557-97BC-4B89-ADB6-D9C93CAAB3DF}"
    with zipfile.ZipFile(path, "r") as source_zip:
        entries = {
            info.filename: source_zip.read(info.filename)
            for info in source_zip.infolist()
        }
        infos = source_zip.infolist()

    ElementTree.register_namespace("", main_ns)
    ElementTree.register_namespace("x14", x14_ns)
    root = ElementTree.fromstring(entries["xl/worksheets/sheet1.xml"])
    extension_list = ElementTree.SubElement(root, f"{{{main_ns}}}extLst")
    extension = ElementTree.SubElement(
        extension_list,
        f"{{{main_ns}}}ext",
        {"uri": extension_uri},
    )
    ElementTree.SubElement(
        extension,
        f"{{{x14_ns}}}dataValidations",
        {"count": "0"},
    )
    entries["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    )

    temporary = path.with_name(f".{path.name}.extension")
    try:
        with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
            for info in infos:
                target_zip.writestr(info, entries[info.filename])
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _add_vba_entry(path: Path):
    with zipfile.ZipFile(path, "a", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/vbaProject.bin", b"not-real-vba")


def test_candidate_appends_filtered_rows_and_preserves_existing_database_rows(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    report_path = tmp_path / "candidate.report.json"
    _source(
        source,
        [
            (101, PROJECT, "already-present"),
            (202, "Other", "filtered-out"),
            (303, PROJECT, "new-row"),
        ],
    )
    _database(database)
    source_before = source.read_bytes()
    database_before = database.read_bytes()

    report = build_candidate(source, database, candidate, report_path=report_path)

    assert source.read_bytes() == source_before
    assert database.read_bytes() == database_before
    assert report.new_ids == (303,)
    assert report.absent_existing_ids == ()
    assert json.loads(report_path.read_text(encoding="utf-8"))["new_ids"] == [303]

    workbook = load_workbook(candidate, data_only=False)
    try:
        slurry = workbook["Slurry"]
        assert [slurry.cell(row, 1).value for row in range(4, 6)] == [101, 303]
        assert slurry["D5"].value == "new-row"

        database_sheet = workbook["db_table"]
        assert database_sheet["A3"].value == 101
        assert [database_sheet.cell(3, column).value for column in range(6, 13)] == [
            f"manual-101-{column}" for column in range(6, 13)
        ]
        assert database_sheet["AD3"].value == "historical-101"
        assert database_sheet["A4"].value == 303
        assert database_sheet["D4"].value == 1
        assert database_sheet["T4"].value == "arbin_sql_h5"
        assert database_sheet["X4"].value == "cycling"
        assert [database_sheet.cell(4, column).value for column in range(6, 13)] == [None] * 7
        assert not any(
            isinstance(database_sheet.cell(4, column).value, str)
            and database_sheet.cell(4, column).value.startswith("=")
            for column in range(1, database_sheet.max_column + 1)
        )
    finally:
        workbook.close()


def test_cellpy_ready_uses_example_formulas_and_actual_row_offsets(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(101, PROJECT, "already-present"), (303, PROJECT, "new-row")])
    _database(database)

    workbook = load_workbook(database)
    try:
        database_sheet = workbook["db_table"]
        database_sheet["A3"] = "=Slurry!A4"
        database_sheet["F3"] = "existing-manual-value"
        workbook.save(database)
    finally:
        workbook.close()

    monkeypatch.setattr(candidate_pipeline, "_recalculate_with_excel", lambda path: None)
    monkeypatch.setattr(candidate_pipeline, "_verify_cellpy_ready", lambda *args, **kwargs: None)

    report = build_candidate(source, database, candidate, cellpy_ready=True)

    assert report.cellpy_ready is True
    assert report.recalculated is True
    workbook = load_workbook(candidate, data_only=False)
    try:
        database_sheet = workbook["db_table"]
        assert database_sheet["A3"].value == "=Slurry!A4"
        assert database_sheet["F3"].value == "existing-manual-value"
        expected_formulas = {
            1: "=Slurry!A5",
            3: "=Slurry!D5",
            15: "=Slurry!AD5",
            16: "=Slurry!AI5",
            17: "=Slurry!AP5",
            18: "=Slurry!F5",
            19: "=Slurry!M5",
            21: "=Slurry!E5",
            22: "=Slurry!C5",
            26: "=Slurry!L5",
            27: "=Z4",
            33: "=Slurry!P5",
        }
        assert {
            column: database_sheet.cell(4, column).value
            for column in expected_formulas
        } == expected_formulas
        assert database_sheet["D4"].value == 1
        assert database_sheet["T4"].value == "arbin_sql_h5"
        assert database_sheet["X4"].value == "cycling"
        approved_columns = {1, 3, 4, 15, 16, 17, 18, 19, 20, 21, 22, 24, 26, 27, 33}
        assert {
            column
            for column in range(1, database_sheet.max_column + 1)
            if database_sheet.cell(4, column).value is not None
        } == approved_columns
    finally:
        workbook.close()


def test_formula_cache_writer_preserves_formula_and_adds_cached_value(tmp_path):
    database = tmp_path / "database.xlsx"
    _database(database)
    workbook = load_workbook(database)
    try:
        database_sheet = workbook["db_table"]
        database_sheet["A3"] = "=Slurry!A4"
        database_sheet["C3"] = "=Slurry!D4"
        workbook.save(database)
    finally:
        workbook.close()

    candidate_pipeline._write_formula_caches(
        database,
        {(3, 1): 101, (3, 3): "existing-value"},
    )

    formulas = load_workbook(database, read_only=True, data_only=False)
    cached = load_workbook(database, read_only=True, data_only=True)
    try:
        assert formulas["db_table"]["A3"].value == "=Slurry!A4"
        assert formulas["db_table"]["C3"].value == "=Slurry!D4"
        assert cached["db_table"]["A3"].value == 101
        assert cached["db_table"]["C3"].value == "existing-value"
    finally:
        formulas.close()
        cached.close()


def test_cellpy_ready_recalculation_failure_leaves_no_candidate_or_input_changes(
    tmp_path, monkeypatch
):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(303, PROJECT, "new-row")])
    _database(database)
    source_before = source.read_bytes()
    database_before = database.read_bytes()

    def fail_recalculation(path):
        raise CandidateBuildError("Excel/COM recalculation unavailable")

    monkeypatch.setattr(candidate_pipeline, "_recalculate_with_excel", fail_recalculation)

    with pytest.raises(CandidateBuildError, match="recalculation unavailable"):
        build_candidate(source, database, candidate, cellpy_ready=True)

    assert not candidate.exists()
    assert source.read_bytes() == source_before
    assert database.read_bytes() == database_before
    assert not list(tmp_path.glob(".candidate.*.xlsx"))


def test_candidate_reports_absent_existing_ids_without_removing_them(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(101, PROJECT, "present")])
    _database(database, ids=(101, 404))

    report = build_candidate(source, database, candidate)

    assert report.absent_existing_ids == (404,)
    workbook = load_workbook(candidate, read_only=True, data_only=False)
    try:
        assert [workbook["Slurry"].cell(row, 1).value for row in range(4, 6)] == [101, 404]
        assert workbook["db_table"]["A4"].value == 404
    finally:
        workbook.close()


def test_noop_candidate_is_byte_identical_and_writes_report(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    report_path = tmp_path / "candidate.report.json"
    _source(source, [(101, PROJECT, "present")])
    _database(database, ids=(101, 404))
    database_before = database.read_bytes()
    save_calls = []
    original_load_workbook = candidate_pipeline.load_workbook

    def load_workbook_with_save_spy(*args, **kwargs):
        workbook = original_load_workbook(*args, **kwargs)
        if not kwargs.get("read_only", False):
            original_save = workbook.save

            def save(*save_args, **save_kwargs):
                save_calls.append(save_args)
                return original_save(*save_args, **save_kwargs)

            workbook.save = save
        return workbook

    monkeypatch.setattr(candidate_pipeline, "load_workbook", load_workbook_with_save_spy)

    report = build_candidate(source, database, candidate, report_path=report_path)

    assert save_calls == []
    assert candidate.read_bytes() == database_before
    assert report.new_ids == ()
    assert json.loads(report_path.read_text(encoding="utf-8")) == {
        "absent_existing_ids": [404],
        "candidate_path": str(candidate.resolve()),
        "cellpy_ready": False,
        "database_path": str(database.resolve()),
        "existing_duplicate_ids": [],
        "existing_slurry_rows": 2,
        "filtered_rows": 1,
        "new_ids": [],
        "recalculated": False,
        "retained_ids": [101],
        "source_path": str(source.resolve()),
        "source_rows": 1,
    }


def test_candidate_accepts_digit_string_ids_in_existing_slurry(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(8231, PROJECT, "existing"), (9000, PROJECT, "new")])
    _database(database, ids=(8231,))
    workbook = load_workbook(database)
    try:
        workbook["Slurry"]["A4"] = "8231"
        workbook.save(database)
    finally:
        workbook.close()

    report = build_candidate(source, database, candidate)

    assert report.retained_ids == (8231,)
    assert report.new_ids == (9000,)
    workbook = load_workbook(candidate, read_only=True, data_only=False)
    try:
        assert workbook["Slurry"]["A4"].value == "8231"
        assert workbook["Slurry"]["A5"].value == 9000
    finally:
        workbook.close()


def test_candidate_appends_slurry_rows_only_to_existing_width(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(9000, PROJECT, "new")])
    workbook = load_workbook(source)
    try:
        workbook["log"].cell(1, 88, "source-only metadata")
        workbook.save(source)
    finally:
        workbook.close()
    _database(database)

    build_candidate(source, database, candidate)

    workbook = load_workbook(candidate, read_only=True, data_only=False)
    try:
        assert workbook["Slurry"].max_column == 87
        assert workbook["Slurry"].cell(5, 88).value is None
    finally:
        workbook.close()


def test_candidate_ignores_invalid_ids_in_filtered_out_projects(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(None, "Other", "ignored"), (0, "Other", "ignored"), (303, PROJECT, "new")])
    _database(database)

    report = build_candidate(source, database, candidate)

    assert report.new_ids == (303,)


def test_candidate_reports_existing_duplicate_ids_without_rewriting_them(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(303, PROJECT, "new")])
    _database(database, ids=(101, 101))

    with pytest.raises(CandidateBuildError, match="duplicate Slurry IDs"):
        build_candidate(source, database, candidate)


def test_candidate_rejects_literal_db_table_id_not_present_in_slurry(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(303, PROJECT, "new")])
    _database(database)
    workbook = load_workbook(database)
    try:
        workbook["db_table"]["A3"] = 999
        workbook.save(database)
    finally:
        workbook.close()

    with pytest.raises(CandidateBuildError, match="orphan db_table literal ID"):
        build_candidate(source, database, candidate)

    assert not candidate.exists()


@pytest.mark.parametrize("bad_target", ["database", "output"])
def test_candidate_rejects_xlsm_database_or_output_before_workbook_access(
    tmp_path, monkeypatch, bad_target
):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    output = tmp_path / "candidate.xlsx"
    _source(source, [(101, PROJECT, "existing")])
    _database(database)
    if bad_target == "database":
        database.rename(tmp_path / "database.xlsm")
        database = tmp_path / "database.xlsm"
    else:
        output = tmp_path / "candidate.xlsm"

    monkeypatch.setattr(
        candidate_pipeline,
        "load_workbook",
        lambda *args, **kwargs: pytest.fail(".xlsm must be rejected before openpyxl access"),
    )
    with pytest.raises(CandidateBuildError, match=r"\.xlsx extension"):
        build_candidate(source, database, output)
    assert not output.exists()


def test_candidate_rejects_vba_disguised_xlsx_before_copy_and_preserves_input(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(101, PROJECT, "existing")])
    _database(database)
    _add_vba_entry(database)
    before = database.read_bytes()
    monkeypatch.setattr(
        candidate_pipeline.shutil,
        "copy2",
        lambda *args, **kwargs: pytest.fail("VBA workbook must be rejected before copy"),
    )

    with pytest.raises(CandidateBuildError, match="contains VBA"):
        build_candidate(source, database, candidate)

    assert database.read_bytes() == before
    assert not candidate.exists()


def _identity_workbook(slurry_ids, database_ids):
    workbook = Workbook()
    slurry = workbook.active
    slurry.title = "Slurry"
    database = workbook.create_sheet("db_table")
    for row, record_id in enumerate(slurry_ids, start=4):
        slurry.cell(row, 1, record_id)
    for row, value in enumerate(database_ids, start=3):
        database.cell(row, 1, value)
    return workbook, slurry, database


@pytest.mark.parametrize(
    ("slurry_ids", "database_ids", "allowed", "message"),
    [
        ((101, 202), ("=Slurry!A4",), (), "missing Slurry IDs"),
        ((101,), (999,), (), "orphan db_table literal ID"),
        ((101,), ("=Slurry!A4", 101), (), "duplicate db_table IDs"),
        ((101,), ("=Slurry!B4",), (), "cannot be validated"),
    ],
    ids=["missing-slurry", "orphan-literal", "duplicate-resolved", "malformed-formula"],
)
def test_database_identity_reconciliation_rejects_invalid_multisets(
    slurry_ids, database_ids, allowed, message
):
    workbook, slurry, database = _identity_workbook(slurry_ids, database_ids)
    try:
        with pytest.raises(CandidateBuildError, match=message):
            candidate_pipeline._database_ids_against_slurry(
                database, slurry, allowed_neware_ids=allowed
            )
    finally:
        workbook.close()


def test_database_identity_reconciliation_accepts_exact_mixed_formula_literal():
    workbook, slurry, database = _identity_workbook((101,), ("='Slurry'!$A$4", 102))
    try:
        assert candidate_pipeline._database_ids_against_slurry(
            database, slurry, allowed_neware_ids=(102,)
        ) == [101, 102]
    finally:
        workbook.close()


def test_publish_outputs_without_manifest_orders_candidate_then_report(tmp_path, monkeypatch):
    candidate_temp = tmp_path / ".candidate.tmp"
    report_temp = tmp_path / ".report.tmp"
    candidate = tmp_path / "candidate.xlsx"
    report = tmp_path / "report.json"
    candidate_temp.write_bytes(b"candidate")
    report_temp.write_bytes(b"report")
    calls = []
    original_replace = candidate_pipeline.os.replace

    def record_replace(source, destination):
        calls.append(Path(destination).name)
        return original_replace(source, destination)

    monkeypatch.setattr(candidate_pipeline.os, "replace", record_replace)
    candidate_pipeline._publish_outputs(
        candidate_temp, candidate, report_temp, report
    )
    assert calls == ["candidate.xlsx", "report.json"]
    assert candidate.read_bytes() == b"candidate"
    assert report.read_bytes() == b"report"


def test_cellpy_ready_noop_recalculates_and_reports_flags(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(101, PROJECT, "present")])
    _database(database)
    calls = []
    monkeypatch.setattr(
        candidate_pipeline,
        "_recalculate_with_excel",
        lambda path: calls.append(("recalculate", path)),
    )
    monkeypatch.setattr(
        candidate_pipeline,
        "_verify_cellpy_ready",
        lambda *args, **kwargs: calls.append(("verify", args[0])),
    )

    report = build_candidate(source, database, candidate, cellpy_ready=True)

    assert report.cellpy_ready is True
    assert report.recalculated is True
    assert [event[0] for event in calls] == ["recalculate", "verify"]


def _formula_database(path: Path):
    _database(path)
    workbook = load_workbook(path)
    try:
        database = workbook["db_table"]
        for column in candidate_pipeline.DB_TABLE_FORMULA_COLUMNS_WITH_ROW_REFERENCE:
            database.cell(3, column, "=Slurry!A4")
        workbook.save(path)
    finally:
        workbook.close()


def _remove_formula_cache(path: Path, coordinate: str):
    with zipfile.ZipFile(path, "r") as source_zip:
        infos = source_zip.infolist()
        entries = {info.filename: source_zip.read(info.filename) for info in infos}
    worksheet = ElementTree.fromstring(entries["xl/worksheets/sheet2.xml"])
    for cell in worksheet.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c"):
        if cell.get("r") == coordinate:
            for child in list(cell):
                if child.tag.rsplit("}", 1)[-1] in {"v", "is"}:
                    cell.remove(child)
    entries["xl/worksheets/sheet2.xml"] = ElementTree.tostring(worksheet, encoding="utf-8")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=path.parent, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
            for info in infos:
                target_zip.writestr(info, entries[info.filename])
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def test_cellpy_ready_validates_existing_formula_caches_and_blank_results(tmp_path):
    database = tmp_path / "database.xlsx"
    _formula_database(database)
    candidate_pipeline._write_formula_caches(
        database,
        {
            (3, column): (101 if column == 1 else None)
            for column in candidate_pipeline.DB_TABLE_FORMULA_COLUMNS_WITH_ROW_REFERENCE
        },
    )

    candidate_pipeline._verify_cellpy_ready(
        database,
        appended_database_rows=(),
        new_ids=(),
    )


def test_cellpy_ready_rejects_missing_existing_formula_cache(tmp_path):
    database = tmp_path / "database.xlsx"
    _formula_database(database)
    candidate_pipeline._write_formula_caches(database, {(3, 1): 101})
    _remove_formula_cache(database, "C3")

    with pytest.raises(CandidateBuildError, match="no cached formula.*column 3"):
        candidate_pipeline._verify_cellpy_ready(
            database,
            appended_database_rows=(),
            new_ids=(),
        )


@pytest.mark.parametrize("bad_id", [candidate_pipeline.EXCEL_SENTINEL, "wrong-type"])
def test_cellpy_ready_rejects_invalid_existing_formula_id_cache(tmp_path, bad_id):
    database = tmp_path / "database.xlsx"
    _formula_database(database)
    candidate_pipeline._write_formula_caches(
        database,
        {
            (3, column): (bad_id if column == 1 else None)
            for column in candidate_pipeline.DB_TABLE_FORMULA_COLUMNS_WITH_ROW_REFERENCE
        },
    )

    with pytest.raises(CandidateBuildError, match="cached ID"):
        candidate_pipeline._verify_cellpy_ready(
            database,
            appended_database_rows=(),
            new_ids=(),
        )


def test_candidate_preserves_representative_supported_workbook_topology(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(101, PROJECT, "present"), (303, PROJECT, "new")])
    _database(database)
    workbook = load_workbook(database)
    try:
        slurry = workbook["Slurry"]
        hidden = workbook.create_sheet("hidden")
        hidden.sheet_state = "hidden"
        slurry.row_dimensions[2].hidden = True
        slurry.merge_cells("B1:C1")
        slurry["A4"].comment = Comment("retain this comment", "test")
        validation = DataValidation(type="whole", operator="between", formula1="1", formula2="999")
        validation.add("A4:A100")
        slurry.add_data_validation(validation)
        slurry.add_table(Table(displayName="SlurryTable", ref="A1:A4"))
        workbook.defined_names.add(DefinedName("SlurryFirst", attr_text="'Slurry'!$A$4"))
        workbook._sheets = [workbook["db_table"], hidden, slurry]
        workbook.save(database)
    finally:
        workbook.close()

    build_candidate(source, database, candidate)

    workbook = load_workbook(candidate, data_only=False)
    try:
        assert workbook.sheetnames == ["db_table", "hidden", "Slurry"]
        assert workbook["hidden"].sheet_state == "hidden"
        assert workbook["Slurry"].row_dimensions[2].hidden is True
        assert "B1:C1" in {str(item) for item in workbook["Slurry"].merged_cells.ranges}
        assert workbook["Slurry"]["A4"].comment.text == "retain this comment"
        assert len(workbook["Slurry"].data_validations.dataValidation) == 1
        assert "SlurryTable" in workbook["Slurry"].tables
        assert workbook.defined_names["SlurryFirst"].attr_text == "'Slurry'!$A$4"
    finally:
        workbook.close()


@pytest.mark.parametrize("bad_id", [None, 0, -1, 1.5, "12", "not-an-id"])
def test_candidate_rejects_invalid_filtered_ids(tmp_path, bad_id):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(bad_id, PROJECT, "bad")])
    _database(database)

    with pytest.raises(CandidateBuildError, match="invalid source ID"):
        build_candidate(source, database, candidate)

    assert not candidate.exists()


def test_candidate_does_not_require_an_unstated_source_header(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(303, PROJECT, "new-row")])
    workbook = load_workbook(source)
    try:
        workbook["log"]["A2"] = "not-required"
        workbook.save(source)
    finally:
        workbook.close()
    _database(database)

    report = build_candidate(source, database, candidate)

    assert report.new_ids == (303,)


def test_candidate_rejects_duplicate_filtered_ids(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(303, PROJECT, "first"), (303, PROJECT, "second")])
    _database(database)

    with pytest.raises(CandidateBuildError, match="duplicate source ID 303"):
        build_candidate(source, database, candidate)

    assert not candidate.exists()


def test_candidate_rejects_in_place_or_production_paths(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    _source(source, [(101, PROJECT, "present")])
    _database(database)

    with pytest.raises(CandidateBuildError, match="separate output"):
        build_candidate(source, database, database)


def test_candidate_rejects_exact_production_database_before_access(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    accesses = []

    def fail_if_accessed(*args, **kwargs):
        accesses.append((args, kwargs))
        raise AssertionError("production workbook must not be accessed")

    monkeypatch.setattr(candidate_pipeline, "load_workbook", fail_if_accessed)
    monkeypatch.setattr(candidate_pipeline.shutil, "copy2", fail_if_accessed)

    with pytest.raises(CandidateBuildError, match="cannot read from or write to production"):
        build_candidate(source, PRODUCTION_DATABASE_PATH, candidate)

    assert accesses == []


def test_candidate_rejects_production_report_path_before_access(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(101, PROJECT, "present")])
    _database(database)

    with pytest.raises(CandidateBuildError, match="cannot read from or write to production"):
        build_candidate(
            source,
            database,
            candidate,
            report_path=PRODUCTION_DATABASE_PATH,
        )

    assert not candidate.exists()


def test_candidate_report_failure_preserves_previous_candidate_report_pair(
    tmp_path, monkeypatch
):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    report_path = tmp_path / "candidate.report.json"
    _source(source, [(101, PROJECT, "present")])
    _database(database)
    build_candidate(source, database, candidate, report_path=report_path)
    candidate_before = candidate.read_bytes()
    report_before = report_path.read_bytes()

    _source(source, [(101, PROJECT, "present"), (303, PROJECT, "new")])
    original_write_text = Path.write_text

    def fail_staged_report(self, *args, **kwargs):
        if self.name.startswith(f".{report_path.name}."):
            raise OSError("injected report write failure")
        return original_write_text(self, *args, **kwargs)

    monkeypatch.setattr(Path, "write_text", fail_staged_report)

    with pytest.raises(OSError, match="injected report write failure"):
        build_candidate(source, database, candidate, report_path=report_path)

    assert candidate.read_bytes() == candidate_before
    assert report_path.read_bytes() == report_before
    assert not list(tmp_path.glob(f".{candidate.name}.*"))
    assert not list(tmp_path.glob(f".{report_path.name}.*"))


def test_default_mode_preserves_cached_values_for_unchanged_formulas(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(101, PROJECT, "present"), (303, PROJECT, "new")])
    _database(database)
    workbook = load_workbook(database)
    try:
        workbook["db_table"]["A3"] = "=Slurry!A4"
        workbook.save(database)
    finally:
        workbook.close()
    candidate_pipeline._write_formula_caches(database, {(3, 1): 101})

    build_candidate(source, database, candidate)

    formulas = load_workbook(candidate, read_only=True, data_only=False)
    cached = load_workbook(candidate, read_only=True, data_only=True)
    try:
        assert formulas["db_table"]["A3"].value == "=Slurry!A4"
        assert cached["db_table"]["A3"].value == 101
    finally:
        formulas.close()
        cached.close()


def test_candidate_rejects_unsupported_data_validation_extension_before_editing(tmp_path):
    source = tmp_path / "source.xlsx"
    database = tmp_path / "database.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _source(source, [(303, PROJECT, "new")])
    _database(database)
    _add_data_validation_extension(database)
    database_before = database.read_bytes()

    with pytest.raises(CandidateBuildError, match="Data Validation extension"):
        build_candidate(source, database, candidate)

    assert not candidate.exists()
    assert database.read_bytes() == database_before


class _FakePipe:
    def __init__(self):
        self.closed = False

    def poll(self, timeout=0):
        return False

    def close(self):
        self.closed = True


class _FakeProcess:
    def __init__(self, target, args, name):
        self.target = target
        self.args = args
        self.name = name
        self.alive = True
        self.terminated = False
        self.exitcode = None
        self.join_calls = []

    def start(self):
        return None

    def join(self, timeout=None):
        self.join_calls.append(timeout)

    def is_alive(self):
        return self.alive

    def terminate(self):
        self.terminated = True
        self.alive = False


class _FakeProcessContext:
    def __init__(self):
        self.process = None

    def Pipe(self, duplex=False):
        return _FakePipe(), _FakePipe()

    def Process(self, target, args, name):
        self.process = _FakeProcess(target, args, name)
        return self.process


def test_excel_recalculation_timeout_terminates_isolated_helper(tmp_path, monkeypatch):
    context = _FakeProcessContext()
    monkeypatch.setattr(candidate_pipeline.os, "name", "nt")
    monkeypatch.setattr(candidate_pipeline.multiprocessing, "get_context", lambda name: context)
    monkeypatch.setattr(candidate_pipeline, "EXCEL_RECALCULATION_TIMEOUT_SECONDS", 0.01)

    with pytest.raises(CandidateBuildError, match="timed out"):
        candidate_pipeline._recalculate_with_excel(tmp_path / "candidate.xlsx")

    assert context.process.terminated is True
    assert context.process.join_calls


def test_excel_recalculation_closes_com_objects_after_failure(tmp_path):
    events = []

    class FakePythoncom:
        def CoInitialize(self):
            events.append("initialize")

        def CoUninitialize(self):
            events.append("uninitialize")

    class FakeWorkbook:
        def Close(self, SaveChanges=False):
            events.append(("close", SaveChanges))

    class FakeWorkbooks:
        def Open(self, **kwargs):
            events.append("open")
            return FakeWorkbook()

    class FakeExcel:
        Workbooks = FakeWorkbooks()

        def __setattr__(self, name, value):
            object.__setattr__(self, name, value)

        def CalculateFullRebuild(self):
            events.append("rebuild")
            raise RuntimeError("injected calculation failure")

        def Quit(self):
            events.append("quit")

    with pytest.raises(CandidateBuildError, match="calculation failure"):
        candidate_pipeline._excel_recalculate_in_process(
            tmp_path / "candidate.xlsx",
            pythoncom_module=FakePythoncom(),
            dispatch_ex=lambda name: FakeExcel(),
        )

    assert events == ["initialize", "open", "rebuild", ("close", False), "quit", "uninitialize"]


def test_missing_pywin32_reports_actionable_cellpy_ready_prerequisite(tmp_path, monkeypatch):
    original_import = builtins.__import__

    def missing_pywin32(name, *args, **kwargs):
        if name in {"pythoncom", "win32com.client"}:
            raise ImportError("pywin32 is unavailable")
        return original_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", missing_pywin32)
    with pytest.raises(CandidateBuildError, match="pywin32>=312"):
        candidate_pipeline._excel_recalculate_in_process(tmp_path / "candidate.xlsx")
