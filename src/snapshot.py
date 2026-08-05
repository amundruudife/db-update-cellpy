"""Atomic values-only local mirror for the hard-coded ``c&p`` sheet."""

import os
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .contracts import LOCAL_SNAPSHOT_PATH, SOURCE_SHEET_NAME
from .source_validation import SourceValidationError, validate_values_only


def snapshot_path(root=None) -> Path:
    """Return the fixed local snapshot path under the workspace root."""

    return Path(root or ".") / LOCAL_SNAPSHOT_PATH


def write_snapshot(values, root=None) -> Path:
    """Validate and atomically write one values-only ``c&p`` workbook."""

    rows = validate_values_only(values)

    destination = snapshot_path(root)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb", suffix=".xlsx", prefix=".Cell_Log_CP.", dir=destination.parent, delete=False
        ) as handle:
            temporary = Path(handle.name)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SOURCE_SHEET_NAME
        for row_number, row in enumerate(rows, start=1):
            for column_number, value in enumerate(row, start=1):
                worksheet.cell(row=row_number, column=column_number, value=value)
        workbook.save(temporary)
        workbook.close()

        reopened = load_workbook(temporary, data_only=False, read_only=True)
        try:
            if reopened.sheetnames != [SOURCE_SHEET_NAME]:
                raise SourceValidationError("Snapshot must contain only the c&p sheet")
            checked_sheet = reopened[SOURCE_SHEET_NAME]
            for row in checked_sheet.iter_rows():
                if any(cell.data_type == "f" for cell in row):
                    raise SourceValidationError("Snapshot contains a formula")
        finally:
            reopened.close()

        os.replace(temporary, destination)
        temporary = None
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)

    return destination
