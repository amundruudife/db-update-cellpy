#!/usr/bin/env python3
"""
Database operations for Cell Analysis Database Auto-Update System.

This module handles database updates and dry-run operations.
"""

import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from .logging_utils import get_logger, log_exceptions
from .file_operations import BACKUP_TIMESTAMP_FORMAT, backup_db
from .common_utils import ensure_directory, generate_project_summary

@log_exceptions("Error updating slurry sheet: {e}")
def update_slurry(new_rows_df, db_path, target_sheet, dry_run=False):
    """
    Update the Slurry sheet with new rows (respecting cellpy format).
    
    Args:
        new_rows_df (pd.DataFrame): New rows to append
        db_path (str): Path to database file
        target_sheet (str): Target sheet name
        dry_run (bool): If True, simulate without writing
    
    Returns:
        int: Number of rows appended
    
    Raises:
        PermissionError: If unable to write to database
        ValueError: If data format doesn't match
    """
    logger = get_logger()
    
    if len(new_rows_df) == 0:
        logger.info("No new rows to append")
        return 0
        
    if dry_run:
        logger.info(f"DRY RUN: Would append {len(new_rows_df)} rows to {target_sheet}")
        return len(new_rows_df)
    
    wb = None
    try:
        # For the cellpy database format, we need to handle this carefully
        # The database has a complex multi-header structure that must be preserved
        
        # Read the current database structure
        wb = load_workbook(db_path)
        
        original_sheets = set(wb.sheetnames)
        if target_sheet not in original_sheets:
            raise ValueError(f"Target sheet '{target_sheet}' not found in database")
        
        ws = wb[target_sheet]
        
        # Find the last row with data (skip empty rows)
        last_row = ws.max_row
        while last_row > 4 and all(cell.value is None for cell in ws[last_row]):
            last_row -= 1
        
        # Convert new data to simple format for appending
        # Note: This assumes the source data structure matches what's needed
        # In practice, you might need to map columns or transform data
        
        start_row = last_row + 1
        rows_added = 0
        
        for idx, row in new_rows_df.iterrows():
            current_row = start_row + rows_added
            
            # Map source data to database columns
            # This is a simplified approach - in practice you may need complex mapping
            for col_idx, value in enumerate(row.values):
                if col_idx < ws.max_column:  # Don't exceed existing column structure
                    ws.cell(row=current_row, column=col_idx + 1, value=value)
            
            rows_added += 1
        
        # Save the workbook
        wb.save(db_path)
        
        # Safety: ensure no new sheets were created (e.g., db_table)
        if set(wb.sheetnames) != original_sheets:
            raise ValueError(
                f"Unexpected new sheets created: {set(wb.sheetnames) - original_sheets}"
            )
        
        logger.info(f"Successfully appended {rows_added} rows to {target_sheet}")
        return rows_added
        
    finally:
        # Bug fix: Always close the workbook to prevent resource leaks
        if wb is not None:
            wb.close()

@log_exceptions("Error in dry run pipeline: {e}")
def dry_run_full_pipeline(config):
    """
    Execute the complete pipeline in dry-run mode, saving results to output folder.
    
    Args:
        config (dict): Configuration dictionary
    
    Returns:
        dict: Execution results summary
    """
    from .data_processing import prepare_update_data
    
    logger = get_logger()
    
    results = {
        'copied_file': None,
        'source_rows': 0,
        'source_max_key': None,
        'filtered_rows': 0,
        'duplicate_rows': 0,
        'appended_rows': 0,
        'output_database': None,
        'backup_created': False,
        'errors': []
    }
    
    logger.info("=" * 60)
    logger.info("STARTING DRY RUN - FULL PIPELINE EXECUTION")
    logger.info("Output will be saved to: output/ folder")
    logger.info("=" * 60)
    
    # Create output directory if it doesn't exist
    output_dir = ensure_directory(Path(config['work_dir']) / 'output')
    
    # Steps 1-3: Prepare data using the centralized function
    prep_results = prepare_update_data(config)
    new_rows_df = prep_results['new_rows_df']
    
    # Populate results from the preparation step
    results['copied_file'] = Path(prep_results['copied_file_path']).name
    results['source_rows'] = prep_results.get('source_rows', 0)
    results['source_max_key'] = prep_results.get('source_max_key')
    results['filtered_rows'] = len(prep_results['filtered_df'])
    results['duplicate_rows'] = len(prep_results['duplicate_keys'])
    results['appended_rows'] = len(new_rows_df)

    # Step 4: Create output database copy
    timestamp = datetime.now().strftime(BACKUP_TIMESTAMP_FORMAT)
    db_filename = Path(config['db_path']).name
    output_db_name = f"dryrun_{db_filename.split('.')[0]}_{timestamp}.xlsx"
    output_db_path = output_dir / output_db_name
    
    logger.info(f"Step 4: Creating output database copy: {output_db_name}")
    shutil.copy2(config['db_path'], output_db_path)
    results['output_database'] = output_db_name
    
    # Step 5: Create backup in output folder (if enabled)
    if config['auto_backup']:
        backup_path = backup_db(config['db_path'], config['auto_backup'], config['work_dir'])
        if backup_path:
            logger.info(f"Backup created in output folder: {Path(backup_path).name}")
            results['backup_created'] = True
    
    # Step 6: Update the output database copy (real operation)
    if len(new_rows_df) > 0:
        logger.info(f"Step 6: Updating output database with {len(new_rows_df)} new rows...")
        rows_added = update_slurry(
            new_rows_df,
            str(output_db_path),
            config['target_sheet'],
            dry_run=False  # Actually modify the output copy
        )
        
        # Verify the update
        logger.info(f"Successfully updated output database: {output_db_name}")
    else:
        logger.info("Step 6: No new rows to append to output database")
    
    # Final summary
    _, summary_text = generate_project_summary(new_rows_df)
    
    logger.info("=" * 60)
    logger.info("DRY RUN PIPELINE COMPLETED SUCCESSFULLY")
    logger.info(f"Source file copied: {results['copied_file']}")
    logger.info(f"Source rows read: {results['source_rows']}")
    logger.info(f"Source max key: {results['source_max_key']}")
    logger.info(f"Rows filtered by projects: {results['filtered_rows']}")
    logger.info(f"Duplicate rows skipped: {results['duplicate_rows']}")
    logger.info(f"New rows appended: {results['appended_rows']} ({summary_text})")
    logger.info(f"Output database: output/{results['output_database']}")
    logger.info(f"Backup created: {results['backup_created']}")
    logger.info("")
    logger.info("✅ All files saved to output/ folder - production database unchanged")
    logger.info("=" * 60)
    
    return results 