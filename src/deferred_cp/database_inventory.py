"""Read-only Gate B inventory and workbook-topology evidence."""

from __future__ import annotations

import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


MANUAL_COLUMNS = frozenset("FGHIJKL")
DATABASE_SHEET = "db_table"
_REF_RE = re.compile(r"(?:=|[^A-Za-z0-9_])'?([^'!]+)'?!\$?[A-Z]+\$?(\d+)")


def classify_cell(value: Any, *, is_formula: bool) -> str:
    if is_formula:
        return "recreate_formula"
    return "retain_existing" if value is not None else "delete"


def compare_cell_inventories(before: Iterable[dict], after: Iterable[dict]) -> list[dict]:
    old = {(c["sheet"], c["coordinate"]): c for c in before}
    new = {(c["sheet"], c["coordinate"]): c for c in after}
    changes = []
    for key in sorted(set(old) | set(new)):
        if key not in new:
            item = {"sheet": key[0], "coordinate": key[1], "classification": "delete", "before": old[key].get("value")}
        elif key not in old:
            item = {"sheet": key[0], "coordinate": key[1], "classification": "recreate", "after": new[key].get("value")}
        elif old[key].get("value") != new[key].get("value"):
            item = {"sheet": key[0], "coordinate": key[1], "classification": "recreate", "before": old[key].get("value"), "after": new[key].get("value")}
        else:
            item = {"sheet": key[0], "coordinate": key[1], "classification": "retain_existing", "before": old[key].get("value")}
        changes.append(item)
    return changes


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _topology(workbook, path: Path) -> dict[str, Any]:
    sheets = {}
    for ws in workbook.worksheets:
        style_ids = Counter()
        comments = []
        validations = []
        formulas = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    style_ids[cell.style_id] += 1
                if cell.data_type == "f":
                    formulas += 1
                if cell.comment:
                    comments.append(cell.coordinate)
        for validation in ws.data_validations.dataValidation:
            validations.append({"type": validation.type, "formula1": validation.formula1, "formula2": validation.formula2, "sqref": str(validation.sqref)})
        sheets[ws.title] = {
            "dimensions": ws.calculate_dimension(),
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "state": ws.sheet_state,
            "styles": {str(k): v for k, v in sorted(style_ids.items())},
            "formulas": formulas,
            "tables": {name: table.ref for name, table in sorted(ws.tables.items())},
            "validations": sorted(validations, key=lambda x: x["sqref"]),
            "comments": sorted(comments),
            "merges": sorted(str(x) for x in ws.merged_cells.ranges),
            "hidden_rows": sorted(str(k) for k, v in ws.row_dimensions.items() if v.hidden),
            "hidden_columns": sorted(str(k) for k, v in ws.column_dimensions.items() if v.hidden),
            "auto_filter": ws.auto_filter.ref,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "column_widths": {str(k): v.width for k, v in sorted(ws.column_dimensions.items()) if v.width is not None},
            "row_heights": {str(k): v.height for k, v in sorted(ws.row_dimensions.items()) if v.height is not None},
            "print_settings": {
                "print_area": str(ws.print_area) if ws.print_area else None,
                "print_title_rows": str(ws.print_title_rows) if ws.print_title_rows else None,
                "print_title_cols": str(ws.print_title_cols) if ws.print_title_cols else None,
                "orientation": ws.page_setup.orientation,
                "paper_size": ws.page_setup.paperSize,
                "fit_to_width": ws.page_setup.fitToWidth,
                "fit_to_height": ws.page_setup.fitToHeight,
            },
        }
    with zipfile.ZipFile(path) as package:
        names = set(package.namelist())
    return {
        "sheet_order": list(workbook.sheetnames),
        "sheets": sheets,
        "defined_names": sorted(str(name) for name in workbook.defined_names),
        "external_links": sorted(name for name in names if name.startswith("xl/externalLinks/")),
        "macros": "xl/vbaProject.bin" in names,
        "calculation": {k: _json_value(v) for k, v in vars(workbook.calculation).items() if v is not None},
    }


def inventory_workbook(path: str | Path, *, row_offset: tuple[str, str, int, int] | None = None, data_start_row: int = 3) -> dict[str, Any]:
    """Capture deterministic evidence without saving or otherwise mutating ``path``."""
    source = Path(path)
    workbook = load_workbook(source, data_only=False, read_only=False, keep_links=True)
    try:
        cells = []
        formula_cells = []
        summary: dict[str, Counter] = defaultdict(Counter)
        ws = workbook[DATABASE_SHEET] if DATABASE_SHEET in workbook.sheetnames else workbook.active
        for row in ws.iter_rows(min_row=data_start_row):
            for cell in row:
                if cell.column_letter in MANUAL_COLUMNS:
                    continue
                if cell.value is None:
                    continue
                is_formula = cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))
                if not is_formula:
                    summary[cell.column_letter]["non_formula_populated"] += 1
                item = {"sheet": ws.title, "coordinate": cell.coordinate, "column": cell.column_letter, "value": _json_value(cell.value), "is_formula": is_formula, "classification": classify_cell(cell.value, is_formula=is_formula)}
                (formula_cells if is_formula else cells).append(item)
        result = _topology(workbook, source)
        result["cells"] = cells
        result["formula_cells"] = formula_cells
        result["column_summary"] = {column: dict(counts) for column, counts in sorted(summary.items())}
        result["row_offset"] = None
        if row_offset:
            target_sheet, source_sheet, target_row, source_row = row_offset
            result["row_offset"] = {"target_sheet": target_sheet, "source_sheet": source_sheet, "target_row": target_row, "source_row": source_row, "offset": source_row - target_row}
        evidence = []
        for cell in formula_cells:
            match = _REF_RE.search(str(cell["value"]))
            if match:
                evidence.append({"target_sheet": cell["sheet"], "target_cell": cell["coordinate"], "source_sheet": match.group(1), "source_row": int(match.group(2)), "target_row": int(re.search(r"\d+$", cell["coordinate"]).group()), "offset": int(match.group(2)) - int(re.search(r"\d+$", cell["coordinate"]).group())})
        result["row_offset_evidence"] = {
            "samples": evidence[:20],
            "observed_offsets": sorted({item["offset"] for item in evidence}),
            "consistent": len({item["offset"] for item in evidence}) <= 1,
            "reference_count": len(evidence),
            "direct_slurry_offsets": sorted({item["offset"] for item in evidence if item["source_sheet"] == "Slurry" and re.fullmatch(r"=Slurry![A-Z]+\d+", str(next(c["value"] for c in formula_cells if c["coordinate"] == item["target_cell"]))) } ),
        }
        return result
    finally:
        workbook.close()


def write_inventory_report(path: str | Path, destination: str | Path, **kwargs) -> dict[str, Any]:
    report = inventory_workbook(path, **kwargs)
    Path(destination).write_text(json.dumps(report, indent=2, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")
    return report
