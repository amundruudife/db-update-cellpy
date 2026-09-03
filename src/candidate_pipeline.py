"""Small candidate-only updater for the retained legacy ``log`` workflow.

This module deliberately has no production replacement capability. It reads a
local legacy Cell Log workbook, keeps the approved project filter, appends only
new IDs to ``Slurry``, and adds minimal ``db_table`` rows by default. The
optional Cellpy-ready mode recreates the example's formula-backed fields.
Existing database rows, including ``b01`` through ``b07``, are never rewritten.
"""

from __future__ import annotations

import json
import gc
import multiprocessing
import os
import posixpath
import re
import shutil
import tempfile
import warnings
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree

from openpyxl import load_workbook

from .contracts import (
    SLURRY_DATA_START_ROW as MIRROR_DATA_START_ROW,
    DATABASE_SHEET_NAME,
    MIRROR_SHEET_NAME,
    SYSTEM_VALUES,
    is_production_database,
)
from .neware import (
    NewareInputError,
    NewareManifest,
    NewareManifestEntry,
    NewareRow,
    NewareSource,
    load_neware_manifest,
    manifest_with_entries,
    read_neware_source,
)


SOURCE_SHEET_NAME = "log"
SOURCE_DATA_START_ROW = 5
SOURCE_PROJECT_COLUMN = 3
SOURCE_KEY_COLUMN = 1
APPROVED_PROJECTS = frozenset(
    {
        "SIS-Larger",
        "SIS-Large",
        "CellMap",
        "Norse-HV",
        "SUMBAT-SP5",
        "SUMBAT",
        "ASAP",
    }
)
DATABASE_DATA_START_ROW = 3
MANUAL_COLUMNS = tuple(range(6, 13))
SYSTEM_COLUMNS = {4: "exists", 20: "instrument", 24: "experiment_type"}
NEWARE_DB_COLUMN_MAP = {
    3: "batch",
    16: "nominal_capacity",
    18: "cell_type",
    22: "project",
    23: "label",
    26: "cell",
    57: "schedule",
    59: "comment_general",
    62: "temperature",
}
NEWARE_SOURCE_BY_TARGET = {
    "label": "Cell ID",
    "batch": "Cell batch",
    "project": "Project",
    "cell_type": "Cell type",
    "cell": "Cell test label",
    "nominal_capacity": "Cell capacity (Ah)",
    "schedule": "Test Schedule",
    "temperature": "Test temp (°C)",
    "comment_general": "Comment",
}
NEWARE_DB_COLUMNS = frozenset(NEWARE_DB_COLUMN_MAP)
DEFAULT_NEWARE_MANIFEST_PATH = Path("source_data") / "neware_id_manifest.json"
DB_TABLE_FORMULA_COLUMNS = {
    1: "A",
    3: "D",
    15: "AD",
    16: "AI",
    17: "AP",
    18: "F",
    19: "M",
    21: "E",
    22: "C",
    26: "L",
    33: "P",
}
DB_TABLE_FORMULA_COLUMNS_WITH_ROW_REFERENCE = (*DB_TABLE_FORMULA_COLUMNS, 27)
EXCEL_ERROR_PREFIX = "#"
EXCEL_SENTINEL = -9223372036854775808


class CandidateBuildError(ValueError):
    """Raised when a candidate cannot be built without guessing or data loss."""


@dataclass(frozen=True)
class CandidateReport:
    """Machine-readable candidate summary.

    ``existing_slurry_rows`` counts nonblank ID-bearing rows, including
    duplicates; it is not a count of unique logical IDs.
    """

    source_path: str
    database_path: str
    candidate_path: str
    source_rows: int
    filtered_rows: int
    existing_slurry_rows: int
    retained_ids: tuple[int, ...]
    new_ids: tuple[int, ...]
    absent_existing_ids: tuple[int, ...]
    existing_duplicate_ids: tuple[int, ...]
    cellpy_ready: bool = False
    recalculated: bool = False
    neware_source_path: str | None = None
    neware_manifest_path: str | None = None
    neware_source_rows: int = 0
    neware_usable_rows: int = 0
    neware_placeholder_rows: tuple[int, ...] = ()
    neware_retained_ids: tuple[int, ...] = ()
    neware_new_ids: tuple[int, ...] = ()
    neware_absent_labels: tuple[str, ...] = ()

    def to_dict(self) -> dict:
        result = asdict(self)
        if self.neware_source_path is None:
            for field in (
                "neware_source_path",
                "neware_manifest_path",
                "neware_source_rows",
                "neware_usable_rows",
                "neware_placeholder_rows",
                "neware_retained_ids",
                "neware_new_ids",
                "neware_absent_labels",
            ):
                result.pop(field, None)
        return result


@dataclass(frozen=True)
class _FormulaCache:
    formula_attributes: tuple[tuple[str, str], ...]
    formula_text: str | None
    cell_type: str | None
    cache_xml: bytes | None


EXCEL_RECALCULATION_TIMEOUT_SECONDS = 120
EXCEL_RECALCULATION_TERMINATION_TIMEOUT_SECONDS = 5
OPENPYXL_EXTENSION_TYPES = {
    "{78C0D931-6437-407D-A8EE-F0AAD7539E65}": "Conditional Formatting",
    "{CCE6A557-97BC-4B89-ADB6-D9C93CAAB3DF}": "Data Validation",
    "{05C60535-1F16-4FD2-B633-F4F36F0B64E0}": "Sparkline Group",
    "{A8765BA9-456A-4DAB-B4F3-ACF838C121DE}": "Slicer List",
    "{FC87AEE6-9EDD-4A0A-B7FB-166176984837}": "Protected Range",
    "{01252117-D84E-4E92-8308-4BE1C098FCBB}": "Ignored Error",
    "{F7C9EE02-42E1-4005-9D12-6889AFFD525C}": "Web Extension",
    "{3A4CF648-6AED-40F4-86FF-DC5316D8AED3}": "Slicer List",
    "{7E03D99C-DC04-49D9-9315-930204A7B6E9}": "Timeline Ref",
}


def _resolved(path: str | Path) -> Path:
    return Path(path).resolve(strict=False)


def _positive_integer(value, *, context: str) -> int:
    if isinstance(value, bool) or value is None:
        raise CandidateBuildError(f"invalid {context}: {value!r}")
    if isinstance(value, int):
        normalized = value
    elif isinstance(value, float) and value.is_integer():
        normalized = int(value)
    else:
        raise CandidateBuildError(f"invalid {context}: {value!r}")
    if normalized <= 0:
        raise CandidateBuildError(f"invalid {context}: {value!r}")
    return normalized


def _nonempty_rows(rows: Iterable[tuple]) -> Iterable[tuple]:
    for row in rows:
        if any(value is not None for value in row):
            yield row


def _read_filtered_source(
    source_path: Path,
    *,
    max_columns: int | None = None,
) -> tuple[int, list[tuple[int, tuple]]]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET_NAME not in workbook.sheetnames:
            raise CandidateBuildError("source workbook must contain the 'log' sheet")
        sheet = workbook[SOURCE_SHEET_NAME]
        source_column_count = max(sheet.max_column, SOURCE_PROJECT_COLUMN)
        column_count = (
            min(source_column_count, max_columns)
            if max_columns is not None
            else source_column_count
        )
        if column_count < SOURCE_PROJECT_COLUMN:
            raise CandidateBuildError("Slurry has no project column")

        source_rows = 0
        filtered = []
        seen = set()
        rows = sheet.iter_rows(
            min_row=SOURCE_DATA_START_ROW,
            max_col=column_count,
            values_only=True,
        )
        for row in _nonempty_rows(rows):
            source_rows += 1
            if row[SOURCE_PROJECT_COLUMN - 1] not in APPROVED_PROJECTS:
                continue
            record_id = _positive_integer(
                row[SOURCE_KEY_COLUMN - 1], context="source ID"
            )
            if record_id in seen:
                raise CandidateBuildError(f"duplicate source ID {record_id}")
            seen.add(record_id)
            filtered.append((record_id, tuple(row)))
        return source_rows, filtered
    finally:
        workbook.close()


def _existing_slurry_id(value, *, context: str) -> int:
    if isinstance(value, str) and value.isascii() and value.isdigit():
        value = int(value)
    return _positive_integer(value, context=context)


def _existing_slurry_ids(sheet) -> tuple[list[int], tuple[int, ...]]:
    ids = []
    for row_number in range(MIRROR_DATA_START_ROW, sheet.max_row + 1):
        value = sheet.cell(row_number, SOURCE_KEY_COLUMN).value
        if value is None:
            continue
        ids.append(
            _existing_slurry_id(value, context=f"existing Slurry ID at A{row_number}")
        )
    duplicate_ids = tuple(
        sorted(record_id for record_id, count in Counter(ids).items() if count > 1)
    )
    return ids, duplicate_ids


_DIRECT_SLURRY_ID_FORMULA = re.compile(
    r"^=(?:'Slurry'|Slurry)!\$?A\$?(\d+)$",
    re.IGNORECASE,
)


def _database_ids_against_slurry(
    database_sheet,
    slurry_sheet,
    *,
    allowed_neware_ids: Iterable[int] = (),
) -> list[int]:
    """Validate db_table IDs against Slurry plus validated Neware IDs."""

    slurry_ids = []
    slurry_ids_by_row = {}
    for row_number in range(MIRROR_DATA_START_ROW, slurry_sheet.max_row + 1):
        value = slurry_sheet.cell(row_number, SOURCE_KEY_COLUMN).value
        if value is None:
            continue
        normalized = _existing_slurry_id(
            value, context=f"existing Slurry ID at A{row_number}"
        )
        slurry_ids.append(normalized)
        slurry_ids_by_row[row_number] = normalized

    slurry_id_set = set(slurry_ids)
    duplicate_slurry_ids = tuple(
        sorted(record_id for record_id, count in Counter(slurry_ids).items() if count > 1)
    )
    if duplicate_slurry_ids:
        raise CandidateBuildError(
            "duplicate Slurry IDs: " + ", ".join(map(str, duplicate_slurry_ids))
        )
    allowed_ids = set(allowed_neware_ids)
    if slurry_id_set & allowed_ids:
        collision = min(slurry_id_set & allowed_ids)
        raise CandidateBuildError(
            f"Neware ID {collision} collides with an existing Slurry ID"
        )

    database_ids = []
    for row_number in range(DATABASE_DATA_START_ROW, database_sheet.max_row + 1):
        row_values = tuple(
            database_sheet.cell(row_number, column).value
            for column in range(1, database_sheet.max_column + 1)
        )
        value = row_values[0]
        if value is None:
            if any(item is not None for item in row_values):
                raise CandidateBuildError(
                    f"db_table row {row_number} has values but no ID in column A"
                )
            continue
        if isinstance(value, str) and value.startswith("="):
            match = _DIRECT_SLURRY_ID_FORMULA.fullmatch(value.strip())
            if match is None:
                raise CandidateBuildError(
                    f"db_table ID formula at A{row_number} cannot be validated against Slurry"
                )
            source_row = int(match.group(1))
            if source_row not in slurry_ids_by_row:
                raise CandidateBuildError(
                    f"db_table ID at A{row_number} references missing Slurry row {source_row}"
                )
            normalized = slurry_ids_by_row[source_row]
        else:
            normalized = _existing_slurry_id(
                value, context=f"existing db_table ID at A{row_number}"
            )
            if normalized not in slurry_id_set and normalized not in allowed_ids:
                raise CandidateBuildError(
                    f"orphan db_table literal ID {normalized} at A{row_number}"
                )
        database_ids.append(normalized)

    duplicate_database_ids = tuple(
        sorted(record_id for record_id, count in Counter(database_ids).items() if count > 1)
    )
    if duplicate_database_ids:
        raise CandidateBuildError(
            "duplicate db_table IDs: " + ", ".join(map(str, duplicate_database_ids))
        )
    expected_slurry_ids = Counter(slurry_ids)
    resolved_ids = Counter(database_ids)
    missing_slurry_ids = expected_slurry_ids - resolved_ids
    if missing_slurry_ids:
        missing = ", ".join(map(str, sorted(missing_slurry_ids.elements())))
        raise CandidateBuildError(f"db_table is missing Slurry IDs: {missing}")
    return database_ids


def _neware_db_row_values(database_sheet, row_number: int) -> dict[str, object]:
    return {
        field: database_sheet.cell(row_number, column).value
        for column, field in NEWARE_DB_COLUMN_MAP.items()
    }


def _validate_existing_neware_rows(
    database_sheet,
    manifest: NewareManifest,
) -> dict[int, int]:
    """Validate manifest-backed literal rows and return ID-to-row locations."""

    rows_by_id: dict[int, int] = {}
    for row_number in range(DATABASE_DATA_START_ROW, database_sheet.max_row + 1):
        value = database_sheet.cell(row_number, 1).value
        if isinstance(value, str) and value.startswith("="):
            continue
        if value is None:
            continue
        record_id = _existing_slurry_id(
            value, context=f"existing db_table ID at A{row_number}"
        )
        entry = next(
            (
                candidate
                for candidate in manifest.entries.values()
                if candidate.record_id == record_id
            ),
            None,
        )
        if entry is None:
            continue
        if record_id in rows_by_id:
            raise CandidateBuildError(
                f"validated Neware ID {record_id} appears more than once in db_table"
            )
        actual = _neware_db_row_values(database_sheet, row_number)
        expected = {
            field: entry.payload[NEWARE_SOURCE_BY_TARGET[field]]
            for field in NEWARE_DB_COLUMN_MAP.values()
        }
        if actual != expected:
            raise CandidateBuildError(
                f"existing Neware payload changed for manifest ID {record_id}"
            )
        if any(
            database_sheet.cell(row_number, column).value != SYSTEM_VALUES[system_name]
            for column, system_name in SYSTEM_COLUMNS.items()
        ):
            raise CandidateBuildError(
                f"existing Neware system values changed for manifest ID {record_id}"
            )
        rows_by_id[record_id] = row_number
    return rows_by_id


def _prepare_neware_records(
    neware_source: NewareSource,
    manifest: NewareManifest,
    *,
    existing_database_ids: Iterable[int],
    existing_slurry_ids: Iterable[int],
    legacy_source_ids: Iterable[int],
) -> tuple[NewareManifest, tuple[tuple[NewareRow, int, bool], ...], tuple[int, ...], tuple[int, ...], tuple[str, ...]]:
    """Validate payload continuity and assign deterministic Neware IDs."""

    existing_database_ids = set(existing_database_ids)
    existing_slurry_ids = set(existing_slurry_ids)
    legacy_source_ids = set(legacy_source_ids)
    manifest_ids = {entry.record_id for entry in manifest.entries.values()}
    if manifest_ids & existing_slurry_ids:
        collision = min(manifest_ids & existing_slurry_ids)
        raise CandidateBuildError(
            f"Neware manifest ID {collision} collides with an existing Slurry ID"
        )
    if manifest_ids & legacy_source_ids:
        collision = min(manifest_ids & legacy_source_ids)
        raise CandidateBuildError(
            f"Neware manifest ID {collision} collides with a legacy source ID"
        )

    entries = dict(manifest.entries)
    assignments: list[tuple[NewareRow, int, bool]] = []
    labels_in_source = set()
    manifest_ahead_labels = {
        label
        for label, entry in entries.items()
        if entry.record_id not in existing_database_ids
    }
    reserved_ids = set(existing_database_ids) | manifest_ids | legacy_source_ids
    next_id = max(reserved_ids, default=0) + 1
    for row in neware_source.rows:
        label = row.cell_test_label
        labels_in_source.add(label)
        existing_entry = entries.get(label)
        if existing_entry is not None:
            if existing_entry.payload != row.payload():
                raise CandidateBuildError(
                    f"changed existing Neware payload for Cell test label {label!r}"
                )
            record_id = existing_entry.record_id
            assignments.append((row, record_id, record_id not in existing_database_ids))
            continue

        while next_id in reserved_ids:
            next_id += 1
        record_id = next_id
        next_id += 1
        reserved_ids.add(record_id)
        entries[label] = NewareManifestEntry(
            record_id=record_id,
            payload=row.payload(),
        )
        assignments.append((row, record_id, True))

    missing_manifest_rows = tuple(sorted(manifest_ahead_labels - labels_in_source))
    if missing_manifest_rows:
        raise CandidateBuildError(
            "manifest-ahead Neware IDs require matching source rows: "
            + ", ".join(repr(label) for label in missing_manifest_rows)
        )

    updated_manifest = manifest_with_entries(manifest, entries)
    new_ids = tuple(record_id for _, record_id, is_new in assignments if is_new)
    retained_ids = tuple(record_id for _, record_id, is_new in assignments if not is_new)
    absent_labels = tuple(sorted(set(entries) - labels_in_source))
    return updated_manifest, tuple(assignments), new_ids, retained_ids, absent_labels


def _assert_safe_paths(
    source_path: Path,
    database_path: Path,
    candidate_path: Path,
    report_path: Path | None,
    manifest_path: Path | None = None,
    neware_source_path: Path | None = None,
) -> None:
    workbook_inputs = ((source_path, "source"), (database_path, "database"))
    if neware_source_path is not None:
        workbook_inputs += ((neware_source_path, "Neware source"),)
    workbook_outputs = ((candidate_path, "candidate"),)
    for path, label in (*workbook_inputs, *workbook_outputs):
        if path.suffix.lower() != ".xlsx":
            raise CandidateBuildError(f"{label} workbook path must use the .xlsx extension")

    input_paths = {source_path, database_path}
    if neware_source_path is not None:
        input_paths.add(neware_source_path)
    if manifest_path is not None:
        input_paths.add(manifest_path)
    if candidate_path in input_paths:
        raise CandidateBuildError("candidate must use a separate output path")
    if report_path is not None and report_path in {*input_paths, candidate_path}:
        raise CandidateBuildError("report must use a separate output path")
    if manifest_path is not None and manifest_path in {source_path, database_path, candidate_path, report_path}:
        raise CandidateBuildError("Neware manifest must use a separate path")
    if any(
        is_production_database(path)
        for path in (
            source_path,
            database_path,
            candidate_path,
            report_path,
            manifest_path,
            neware_source_path,
        )
        if path is not None
    ):
        raise CandidateBuildError("candidate workflow cannot read from or write to production")
    if not source_path.is_file():
        raise FileNotFoundError(f"source workbook not found: {source_path}")
    if not database_path.is_file():
        raise FileNotFoundError(f"database workbook not found: {database_path}")
    if neware_source_path is not None and not neware_source_path.is_file():
        raise FileNotFoundError(f"Neware source workbook not found: {neware_source_path}")

    for path, label in workbook_inputs:
        try:
            with zipfile.ZipFile(path, "r") as source_zip:
                if "xl/vbaProject.bin" in source_zip.namelist():
                    raise CandidateBuildError(
                        f"{label} workbook contains VBA; .xlsm/VBA workbooks are rejected"
                    )
        except CandidateBuildError:
            raise
        except (OSError, zipfile.BadZipFile) as exc:
            raise CandidateBuildError(f"{label} workbook is not a valid .xlsx OOXML file") from exc
    if candidate_path.exists():
        try:
            with zipfile.ZipFile(candidate_path, "r") as candidate_zip:
                if "xl/vbaProject.bin" in candidate_zip.namelist():
                    raise CandidateBuildError(
                        "candidate workbook contains VBA; .xlsm/VBA workbooks are rejected"
                    )
        except CandidateBuildError:
            raise
        except (OSError, zipfile.BadZipFile) as exc:
            raise CandidateBuildError("candidate output is not a valid .xlsx OOXML file") from exc


def _worksheet_values(sheet, *, max_row: int, max_column: int) -> tuple[tuple, ...]:
    return tuple(
        tuple(row)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_column,
            values_only=True,
        )
    )


def _is_excel_error(value) -> bool:
    return (
        isinstance(value, str) and value.startswith(EXCEL_ERROR_PREFIX)
    ) or (isinstance(value, int) and value < -1_000_000)


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _assert_openpyxl_safe_workbook(path: Path) -> None:
    """Reject worksheet extensions that OpenPyXL would silently remove."""

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    try:
        with zipfile.ZipFile(path, "r") as source_zip:
            worksheet_entries = tuple(
                info.filename
                for info in source_zip.infolist()
                if info.filename.startswith("xl/worksheets/")
                and info.filename.endswith(".xml")
            )
            for worksheet_name in worksheet_entries:
                root = ElementTree.fromstring(source_zip.read(worksheet_name))
                for extension_list in root.iter(f"{{{main_ns}}}extLst"):
                    for extension in extension_list:
                        uri = (extension.get("uri") or "").upper()
                        extension_name = OPENPYXL_EXTENSION_TYPES.get(uri, "Unknown")
                        if any(
                            _local_name(child.tag) == "dataValidations"
                            for child in extension.iter()
                        ):
                            extension_name = "Data Validation"
                        raise CandidateBuildError(
                            f"OpenPyXL would remove unsupported {extension_name} "
                            f"extension in {worksheet_name}"
                        )
    except CandidateBuildError:
        raise
    except (OSError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        raise CandidateBuildError(
            f"could not inspect workbook features before editing: {path}"
        ) from exc


def _load_workbook_for_edit(path: Path):
    """Load a workbook only when OpenPyXL reports no lossy feature handling."""

    _assert_openpyxl_safe_workbook(path)
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        workbook = load_workbook(path, data_only=False)
    unsupported = [warning for warning in caught if issubclass(warning.category, UserWarning)]
    if unsupported:
        workbook.close()
        message = str(unsupported[0].message)
        raise CandidateBuildError(f"OpenPyXL reported an unsupported workbook feature: {message}")
    return workbook


def _save_workbook_for_edit(workbook, path: Path) -> None:
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        workbook.save(path)
    unsupported = [warning for warning in caught if issubclass(warning.category, UserWarning)]
    if unsupported:
        raise CandidateBuildError(
            f"OpenPyXL reported an unsupported workbook feature while saving: "
            f"{unsupported[0].message}"
        )


def _formula_cells(candidate_path: Path) -> tuple[tuple[int, int, str], ...]:
    workbook = load_workbook(candidate_path, read_only=True, data_only=False)
    try:
        sheet = workbook[DATABASE_SHEET_NAME]
        return tuple(
            (cell.row, cell.column, cell.coordinate)
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
    finally:
        workbook.close()


def _formula_cache_snapshot(candidate_path: Path) -> dict[tuple[str, str], _FormulaCache]:
    """Capture formula XML and cached values before OpenPyXL edits the package."""

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    snapshot = {}
    try:
        with zipfile.ZipFile(candidate_path, "r") as source_zip:
            worksheet_entries = tuple(
                info.filename
                for info in source_zip.infolist()
                if info.filename.startswith("xl/worksheets/")
                and info.filename.endswith(".xml")
            )
            for worksheet_name in worksheet_entries:
                root = ElementTree.fromstring(source_zip.read(worksheet_name))
                for cell in root.iter(f"{{{main_ns}}}c"):
                    coordinate = cell.get("r")
                    formula = cell.find(f"{{{main_ns}}}f")
                    if coordinate is None or formula is None:
                        continue
                    cache_elements = [
                        child
                        for child in list(cell)
                        if _local_name(child.tag) in {"v", "is"}
                    ]
                    unsupported_children = [
                        child
                        for child in list(cell)
                        if child is not formula and _local_name(child.tag) not in {"v", "is"}
                    ]
                    if len(cache_elements) > 1 or unsupported_children:
                        raise CandidateBuildError(
                            f"cannot safely preserve formula cache at "
                            f"{worksheet_name}!{coordinate}"
                        )
                    cache_xml = (
                        ElementTree.tostring(cache_elements[0], encoding="utf-8")
                        if cache_elements
                        else None
                    )
                    snapshot[(worksheet_name, coordinate)] = _FormulaCache(
                        formula_attributes=tuple(sorted(formula.attrib.items())),
                        formula_text=formula.text,
                        cell_type=cell.get("t"),
                        cache_xml=cache_xml,
                    )
    except CandidateBuildError:
        raise
    except (OSError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        raise CandidateBuildError(
            f"could not capture formula caches before editing: {candidate_path}"
        ) from exc
    return snapshot


def _rewrite_zip_entries(
    candidate_path: Path,
    entries: dict[str, bytes],
    infos: Iterable[zipfile.ZipInfo],
    *,
    prefix: str,
) -> None:
    temporary = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            suffix=".xlsx",
            prefix=prefix,
            dir=candidate_path.parent,
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
        with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
            for info in infos:
                target_zip.writestr(info, entries[info.filename])
        os.replace(temporary, candidate_path)
        temporary = None
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def _restore_formula_caches(
    candidate_path: Path,
    snapshot: dict[tuple[str, str], _FormulaCache],
) -> None:
    """Restore original caches while rejecting any formula-text change."""

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    entries = {}
    with zipfile.ZipFile(candidate_path, "r") as source_zip:
        infos = source_zip.infolist()
        for info in infos:
            entries[info.filename] = source_zip.read(info.filename)

    roots = {}
    for worksheet_name, coordinate in snapshot:
        if worksheet_name not in roots:
            try:
                roots[worksheet_name] = ElementTree.fromstring(entries[worksheet_name])
            except (KeyError, ElementTree.ParseError) as exc:
                raise CandidateBuildError(
                    f"could not safely restore formula cache package entry {worksheet_name}"
                ) from exc
        root = roots[worksheet_name]
        cells = {
            cell.get("r"): cell
            for cell in root.iter(f"{{{main_ns}}}c")
            if cell.get("r")
        }
        cell = cells.get(coordinate)
        expected = snapshot[(worksheet_name, coordinate)]
        if cell is None:
            raise CandidateBuildError(
                f"formula cell is missing while restoring cache: "
                f"{worksheet_name}!{coordinate}"
            )
        formula = cell.find(f"{{{main_ns}}}f")
        if (
            formula is None
            or tuple(sorted(formula.attrib.items())) != expected.formula_attributes
            or formula.text != expected.formula_text
        ):
            raise CandidateBuildError(
                f"formula text changed; refusing to restore cache at "
                f"{worksheet_name}!{coordinate}"
            )
        if expected.cell_type is None:
            cell.attrib.pop("t", None)
        else:
            cell.set("t", expected.cell_type)
        for child in list(cell):
            if _local_name(child.tag) in {"v", "is"}:
                cell.remove(child)
        if expected.cache_xml is not None:
            cell.append(ElementTree.fromstring(expected.cache_xml))

    for worksheet_name, root in roots.items():
        entries[worksheet_name] = ElementTree.tostring(root, encoding="utf-8")
    _rewrite_zip_entries(
        candidate_path,
        entries,
        infos,
        prefix=f".{candidate_path.stem}.formula-cache-",
    )


def _database_formula_caches(candidate_path: Path) -> dict[str, _FormulaCache]:
    with zipfile.ZipFile(candidate_path, "r") as source_zip:
        entries = {
            info.filename: source_zip.read(info.filename)
            for info in source_zip.infolist()
        }
    worksheet_name = _worksheet_part(entries, DATABASE_SHEET_NAME)
    snapshot = _formula_cache_snapshot(candidate_path)
    return {
        coordinate: cache
        for (part, coordinate), cache in snapshot.items()
        if part == worksheet_name
    }


def _excel_formula_values(sheet, formula_cells) -> dict[tuple[int, int], object]:
    used_range = sheet.UsedRange
    first_row = int(used_range.Row)
    first_column = int(used_range.Column)
    raw_values = used_range.Value2
    if isinstance(raw_values, tuple):
        if raw_values and isinstance(raw_values[0], tuple):
            values = raw_values
        else:
            values = (raw_values,)
    else:
        values = ((raw_values,),)

    result = {}
    for row_number, column_number, coordinate in formula_cells:
        row_index = row_number - first_row
        column_index = column_number - first_column
        if row_index < 0 or column_index < 0:
            raise CandidateBuildError(f"Excel formula cell is outside UsedRange: {coordinate}")
        try:
            value = values[row_index][column_index]
        except (IndexError, TypeError) as exc:
            raise CandidateBuildError(
                f"Excel did not return a value for formula cell {coordinate}"
            ) from exc
        if _is_excel_error(value):
            raise CandidateBuildError(f"Excel returned an error for formula cell {coordinate}")
        result[(row_number, column_number)] = value
    return result


def _worksheet_part(entries: dict[str, bytes], sheet_name: str) -> str:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    document_relationship_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_relationship_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    workbook_root = ElementTree.fromstring(entries["xl/workbook.xml"])
    sheet = next(
        (
            candidate
            for candidate in workbook_root.findall(f"{{{main_ns}}}sheets/{{{main_ns}}}sheet")
            if candidate.get("name") == sheet_name
        ),
        None,
    )
    if sheet is None:
        raise CandidateBuildError(f"workbook is missing the {sheet_name} sheet")
    relationship_id = sheet.get(f"{{{document_relationship_ns}}}id")
    relationships_root = ElementTree.fromstring(entries["xl/_rels/workbook.xml.rels"])
    relationship = next(
        (
            candidate
            for candidate in relationships_root.findall(f"{{{package_relationship_ns}}}Relationship")
            if candidate.get("Id") == relationship_id
        ),
        None,
    )
    if relationship is None:
        raise CandidateBuildError(f"workbook relationship is missing for {sheet_name}")
    target = relationship.get("Target")
    if target is None:
        raise CandidateBuildError(f"workbook relationship has no target for {sheet_name}")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _write_formula_caches(
    candidate_path: Path,
    formula_values: dict[tuple[int, int], object],
) -> None:
    """Persist values evaluated by Excel without replacing the formulas."""

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    entries = {}
    with zipfile.ZipFile(candidate_path, "r") as source_zip:
        infos = source_zip.infolist()
        for info in infos:
            entries[info.filename] = source_zip.read(info.filename)

    worksheet_name = _worksheet_part(entries, DATABASE_SHEET_NAME)
    worksheet_root = ElementTree.fromstring(entries[worksheet_name])
    cells = {
        cell.get("r"): cell
        for cell in worksheet_root.iter(f"{{{main_ns}}}c")
        if cell.get("r")
    }
    for (row_number, column_number), value in formula_values.items():
        from openpyxl.utils import get_column_letter

        coordinate = f"{get_column_letter(column_number)}{row_number}"
        cell = cells.get(coordinate)
        if cell is None:
            raise CandidateBuildError(f"formula cell is missing from workbook XML: {coordinate}")
        cached = cell.find(f"{{{main_ns}}}v")
        if cached is None:
            cached = ElementTree.SubElement(cell, f"{{{main_ns}}}v")
        if value is None:
            cell.attrib.pop("t", None)
            cached.text = None
        elif isinstance(value, bool):
            cell.set("t", "b")
            cached.text = "1" if value else "0"
        elif isinstance(value, (int, float)):
            cell.attrib.pop("t", None)
            cached.text = repr(value)
        else:
            if _is_excel_error(value):
                raise CandidateBuildError(f"Excel returned an error for formula cell {coordinate}")
            cell.set("t", "str")
            cached.text = str(value)
    entries[worksheet_name] = ElementTree.tostring(worksheet_root, encoding="utf-8")

    _rewrite_zip_entries(
        candidate_path,
        entries,
        infos,
        prefix=f".{candidate_path.stem}.cache-",
    )


def _excel_recalculate_in_process(
    candidate_path: Path,
    *,
    pythoncom_module=None,
    dispatch_ex=None,
) -> None:
    """Run one bounded unit of Excel work with deterministic COM cleanup."""

    if pythoncom_module is None or dispatch_ex is None:
        try:
            import pythoncom as imported_pythoncom
            from win32com.client import DispatchEx as imported_dispatch_ex
        except Exception as exc:  # pragma: no cover - depends on the operator runtime
            raise CandidateBuildError(
                "Excel/COM recalculation requires pywin32>=312; install pywin32 on Windows"
            ) from exc
        pythoncom_module = pythoncom_module or imported_pythoncom
        dispatch_ex = dispatch_ex or imported_dispatch_ex

    excel = None
    workbook = None
    com_initialized = False
    failure = None
    cleanup_errors = []
    formula_values = None
    try:
        pythoncom_module.CoInitialize()
        com_initialized = True
        formula_cells = _formula_cells(candidate_path) if candidate_path.is_file() else ()
        excel = dispatch_ex("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False
        excel.AutomationSecurity = 3
        try:
            excel.Calculation = -4105  # xlCalculationAutomatic
        except Exception:
            pass
        workbook = excel.Workbooks.Open(
            Filename=str(candidate_path),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )
        excel.CalculateFullRebuild()
        formula_values = _excel_formula_values(
            workbook.Worksheets(DATABASE_SHEET_NAME), formula_cells
        )
        workbook.Save()
    except CandidateBuildError as exc:
        failure = exc
    except Exception as exc:  # pragma: no cover - depends on the operator runtime
        failure = CandidateBuildError(f"Excel/COM recalculation failed: {exc}")
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception as exc:  # pragma: no cover - depends on the operator runtime
                cleanup_errors.append(f"workbook close failed: {exc}")
        if excel is not None:
            try:
                excel.Quit()
            except Exception as exc:  # pragma: no cover - depends on the operator runtime
                cleanup_errors.append(f"Excel quit failed: {exc}")
        workbook = None
        excel = None
        gc.collect()
        if com_initialized:
            try:
                pythoncom_module.CoUninitialize()
            except Exception as exc:  # pragma: no cover - depends on the operator runtime
                cleanup_errors.append(f"COM uninitialization failed: {exc}")

    if failure is None and cleanup_errors:
        failure = CandidateBuildError("Excel/COM cleanup failed: " + "; ".join(cleanup_errors))
    if failure is not None:
        raise failure
    if formula_values is not None:
        try:
            _write_formula_caches(candidate_path, formula_values)
        except CandidateBuildError:
            raise
        except Exception as exc:
            raise CandidateBuildError(f"could not persist Excel formula caches: {exc}") from exc


def _excel_recalculate_worker(candidate_path: str, connection) -> None:
    try:
        _excel_recalculate_in_process(Path(candidate_path))
    except BaseException as exc:  # pragma: no cover - executed in a child process
        connection.send(("error", str(exc)))
    else:  # pragma: no cover - executed in a child process
        connection.send(("ok", ""))
    finally:  # pragma: no cover - executed in a child process
        connection.close()


def _recalculate_with_excel(candidate_path: Path) -> None:
    """Run Excel in an isolated process with a hard parent-side timeout."""

    if os.name != "nt":
        raise CandidateBuildError("Excel/COM recalculation is available only on Windows")
    try:
        context = multiprocessing.get_context("spawn")
        receive, send = context.Pipe(duplex=False)
        process = context.Process(
            target=_excel_recalculate_worker,
            args=(str(candidate_path), send),
            name="candidate-excel-recalculation",
        )
        process.start()
        process.join(EXCEL_RECALCULATION_TIMEOUT_SECONDS)
        if process.is_alive():
            process.terminate()
            process.join(EXCEL_RECALCULATION_TERMINATION_TIMEOUT_SECONDS)
            raise CandidateBuildError(
                f"Excel/COM recalculation timed out after "
                f"{EXCEL_RECALCULATION_TIMEOUT_SECONDS} seconds"
            )
        if receive.poll(0.1):
            status, message = receive.recv()
            if status == "error":
                raise CandidateBuildError(message or "Excel/COM recalculation failed")
        elif process.exitcode:
            raise CandidateBuildError(
                f"Excel/COM recalculation helper exited with code {process.exitcode}"
            )
        else:
            raise CandidateBuildError("Excel/COM recalculation helper returned no result")
    except CandidateBuildError:
        raise
    except Exception as exc:  # pragma: no cover - depends on the operator runtime
        raise CandidateBuildError(f"Excel/COM recalculation unavailable: {exc}") from exc
    finally:
        for connection in (locals().get("receive"), locals().get("send")):
            if connection is not None:
                try:
                    connection.close()
                except Exception:
                    pass


def _verify_cellpy_ready(
    candidate_path: Path,
    *,
    appended_database_rows: tuple[int, ...],
    new_ids: tuple[int, ...],
) -> None:
    """Verify cached formula values needed by Cellpy's database reader."""

    formulas = load_workbook(candidate_path, read_only=True, data_only=False)
    cached = load_workbook(candidate_path, read_only=True, data_only=True)
    try:
        formula_sheet = formulas[DATABASE_SHEET_NAME]
        cached_sheet = cached[DATABASE_SHEET_NAME]
        formula_caches = _database_formula_caches(candidate_path)

        appended_ids_by_row = dict(zip(appended_database_rows, new_ids))
        cached_new_ids = []
        max_column = max(
            formula_sheet.max_column,
            max(DB_TABLE_FORMULA_COLUMNS_WITH_ROW_REFERENCE),
        )
        formula_rows = formula_sheet.iter_rows(
            min_row=DATABASE_DATA_START_ROW,
            max_row=formula_sheet.max_row,
            max_col=max_column,
        )
        cached_rows = cached_sheet.iter_rows(
            min_row=DATABASE_DATA_START_ROW,
            max_row=formula_sheet.max_row,
            max_col=max_column,
        )
        for formula_row, cached_row in zip(formula_rows, cached_rows):
            row_number = formula_row[0].row
            formula_id = formula_row[0].value
            cached_id = cached_row[0].value
            if isinstance(formula_id, str) and formula_id.startswith("="):
                id_cache = formula_caches.get(formula_row[0].coordinate)
                if id_cache is None or not (id_cache.formula_text or "").strip():
                    raise CandidateBuildError(
                        f"Cellpy-ready candidate has a malformed formula at db_table A{row_number}"
                    )
                if cached_id is None or cached_id == EXCEL_SENTINEL or _is_excel_error(cached_id):
                    raise CandidateBuildError(
                        f"Cellpy-ready candidate has no cached ID at db_table A{row_number}"
                    )
                try:
                    _existing_slurry_id(
                        cached_id,
                        context=f"cached db_table ID at A{row_number}",
                    )
                except CandidateBuildError as exc:
                    raise CandidateBuildError(
                        f"Cellpy-ready cached ID invalid at db_table A{row_number}: {exc}"
                    ) from exc

            required_formulas = {
                column: formula_row[column - 1].value
                for column in DB_TABLE_FORMULA_COLUMNS_WITH_ROW_REFERENCE
            }
            row_has_required_formula = any(
                isinstance(formula, str) and formula.startswith("=")
                for formula in required_formulas.values()
            )
            expected_id = appended_ids_by_row.get(row_number)
            if not row_has_required_formula and expected_id is None:
                continue
            for column, formula in required_formulas.items():
                value = cached_row[column - 1].value
                if not isinstance(formula, str) or not formula.startswith("="):
                    raise CandidateBuildError(
                        f"Cellpy-ready candidate is missing formula at "
                        f"db_table row {row_number}, column {column}"
                    )
                cache = formula_caches.get(formula_row[column - 1].coordinate)
                if cache is None or cache.cache_xml is None:
                    raise CandidateBuildError(
                        f"Cellpy-ready candidate has no cached formula at "
                        f"db_table row {row_number}, column {column}"
                    )
                if not (cache.formula_text or "").strip():
                    raise CandidateBuildError(
                        f"Cellpy-ready candidate has a malformed formula at "
                        f"db_table row {row_number}, column {column}"
                    )
                if (
                    value == EXCEL_SENTINEL
                    or _is_excel_error(value)
                    or EXCEL_ERROR_PREFIX in formula
                ):
                    raise CandidateBuildError(
                        f"Cellpy-ready candidate has an invalid cached formula at "
                        f"db_table row {row_number}, column {column}"
                    )
            if expected_id is None:
                continue
            try:
                normalized_id = _existing_slurry_id(
                    cached_id, context=f"cached db_table ID at A{row_number}"
                )
            except CandidateBuildError as exc:
                raise CandidateBuildError(
                    f"Cellpy-ready cached ID invalid at db_table A{row_number}: {exc}"
                ) from exc
            if normalized_id != expected_id:
                raise CandidateBuildError(
                    f"Cellpy-ready cached ID mismatch at db_table A{row_number}"
                )
            cached_new_ids.append(normalized_id)

        if len(cached_new_ids) != len(appended_database_rows):
            raise CandidateBuildError("Cellpy-ready candidate is missing appended rows")

        if len(cached_new_ids) != len(set(cached_new_ids)):
            raise CandidateBuildError("Cellpy-ready candidate contains duplicate new IDs")
    finally:
        formulas.close()
        cached.close()


def _verify_candidate(
    candidate_path: Path,
    *,
    initial_slurry_rows: int,
    initial_slurry_columns: int,
    existing_slurry_values: tuple[tuple, ...],
    initial_database_rows: int,
    initial_database_columns: int,
    existing_database_values: tuple[tuple, ...],
    manual_values: tuple[tuple, ...],
    appended_slurry_rows: tuple[int, ...],
    appended_database_rows: tuple[int, ...],
    new_ids: tuple[int, ...],
    neware_appended_rows: tuple[tuple[int, int, NewareRow], ...],
    cellpy_ready: bool,
) -> None:
    workbook = load_workbook(candidate_path, read_only=True, data_only=False)
    try:
        if MIRROR_SHEET_NAME not in workbook.sheetnames or DATABASE_SHEET_NAME not in workbook.sheetnames:
            raise CandidateBuildError("candidate is missing Slurry or db_table")
        slurry = workbook[MIRROR_SHEET_NAME]
        database = workbook[DATABASE_SHEET_NAME]

        actual_slurry_values = _worksheet_values(
            slurry,
            max_row=initial_slurry_rows,
            max_column=initial_slurry_columns,
        )
        if actual_slurry_values != existing_slurry_values:
            raise CandidateBuildError("candidate changed existing Slurry rows")

        actual_slurry_ids = tuple(
            _positive_integer(
                slurry.cell(row_number, SOURCE_KEY_COLUMN).value,
                context=f"candidate Slurry ID at A{row_number}",
            )
            for row_number in appended_slurry_rows
        )
        if actual_slurry_ids != new_ids:
            raise CandidateBuildError("candidate Slurry append verification failed")

        actual_database_values = _worksheet_values(
            database,
            max_row=initial_database_rows,
            max_column=initial_database_columns,
        )
        if actual_database_values != existing_database_values:
            raise CandidateBuildError("candidate changed existing db_table rows")

        manual_start = MANUAL_COLUMNS[0] - 1
        actual_manual = tuple(
            row[manual_start : manual_start + len(MANUAL_COLUMNS)]
            for row in actual_database_values[DATABASE_DATA_START_ROW - 1 :]
        )
        if actual_manual != manual_values:
            raise CandidateBuildError("candidate changed existing b01:b07 values")

        if len(appended_database_rows) != len(new_ids):
            raise CandidateBuildError("candidate db_table append verification failed")
        for row_number, record_id in zip(appended_database_rows, new_ids):
            if cellpy_ready:
                if not (
                    isinstance(database.cell(row_number, 1).value, str)
                    and database.cell(row_number, 1).value.startswith("=")
                ):
                    raise CandidateBuildError("candidate db_table formula verification failed")
            elif database.cell(row_number, 1).value != record_id:
                raise CandidateBuildError("candidate db_table ID verification failed")
            allowed_columns = {1, *SYSTEM_COLUMNS}
            if cellpy_ready:
                allowed_columns.update(DB_TABLE_FORMULA_COLUMNS_WITH_ROW_REFERENCE)
            if any(
                column not in allowed_columns and database.cell(row_number, column).value is not None
                for column in range(1, database.max_column + 1)
            ):
                raise CandidateBuildError("new db_table cells must be blank except approved system values")
            for column, system_name in SYSTEM_COLUMNS.items():
                if database.cell(row_number, column).value != SYSTEM_VALUES[system_name]:
                    raise CandidateBuildError("candidate system-value verification failed")

        for row_number, record_id, neware_row in neware_appended_rows:
            if database.cell(row_number, 1).value != record_id:
                raise CandidateBuildError("candidate Neware db_table ID verification failed")
            allowed_columns = {1, *SYSTEM_COLUMNS, *NEWARE_DB_COLUMNS}
            if any(
                column not in allowed_columns and database.cell(row_number, column).value is not None
                for column in range(1, database.max_column + 1)
            ):
                raise CandidateBuildError(
                    "new Neware db_table cells must be blank except approved mappings"
                )
            mapped_values = neware_row.mapped_values()
            for column, field in NEWARE_DB_COLUMN_MAP.items():
                if database.cell(row_number, column).value != mapped_values[field]:
                    raise CandidateBuildError(
                        f"candidate Neware mapping verification failed at db_table row {row_number}"
                    )
            for column, system_name in SYSTEM_COLUMNS.items():
                if database.cell(row_number, column).value != SYSTEM_VALUES[system_name]:
                    raise CandidateBuildError("candidate Neware system-value verification failed")
    finally:
        workbook.close()


def _reserve_output_path(path: Path, *, label: str) -> Path:
    handle = tempfile.NamedTemporaryFile(
        mode="wb",
        suffix=".tmp",
        prefix=f".{path.name}.{label}-",
        dir=path.parent,
        delete=False,
    )
    reserved = Path(handle.name)
    handle.close()
    reserved.unlink(missing_ok=True)
    return reserved


def _stage_report(report_path: Path, result: CandidateReport) -> Path:
    staged = _reserve_output_path(report_path, label="report")
    try:
        staged.write_text(
            json.dumps(result.to_dict(), indent=2, sort_keys=True),
            encoding="utf-8",
        )
    except BaseException:
        staged.unlink(missing_ok=True)
        raise
    return staged


def _stage_manifest(manifest_path: Path, manifest: NewareManifest) -> Path:
    staged = _reserve_output_path(manifest_path, label="manifest")
    try:
        staged.write_text(
            json.dumps(manifest.document(), indent=2, sort_keys=True),
            encoding="utf-8",
        )
    except BaseException:
        staged.unlink(missing_ok=True)
        raise
    return staged


def _publish_outputs(
    candidate_temporary: Path,
    candidate_path: Path,
    report_temporary: Path | None,
    report_path: Path | None,
    manifest_temporary: Path | None = None,
    manifest_path: Path | None = None,
) -> None:
    """Publish staged artifacts in restart-recoverable replacement order.

    Each replacement is independently durable. An interruption may leave a
    mixture of old and new artifacts; rerunning reuses manifest IDs and
    deterministically completes the publication.
    """

    ordered = []
    if manifest_temporary is not None and manifest_path is not None:
        ordered.append((manifest_temporary, manifest_path))
    ordered.append((candidate_temporary, candidate_path))
    if report_temporary is not None and report_path is not None:
        ordered.append((report_temporary, report_path))
    for temporary, destination in ordered:
        os.replace(temporary, destination)


def build_candidate(
    source_path: str | Path,
    database_path: str | Path,
    candidate_path: str | Path,
    *,
    report_path: str | Path | None = None,
    cellpy_ready: bool = False,
    neware_source_path: str | Path | None = None,
    neware_manifest_path: str | Path | None = None,
) -> CandidateReport:
    """Build and verify a candidate without touching source or production."""

    source = _resolved(source_path)
    database = _resolved(database_path)
    candidate = _resolved(candidate_path)
    report = _resolved(report_path) if report_path is not None else None
    neware_source_path = (
        _resolved(neware_source_path) if neware_source_path is not None else None
    )
    manifest_path = None
    if neware_source_path is not None or neware_manifest_path is not None:
        manifest_path = _resolved(
            neware_manifest_path
            if neware_manifest_path is not None
            else DEFAULT_NEWARE_MANIFEST_PATH
        )
    if cellpy_ready and neware_source_path is not None:
        raise CandidateBuildError(
            "--cellpy-ready cannot be combined with --neware-source until a Neware formula mapping is approved"
        )
    _assert_safe_paths(
        source,
        database,
        candidate,
        report,
        manifest_path,
        neware_source_path,
    )

    candidate.parent.mkdir(parents=True, exist_ok=True)
    if report is not None:
        report.parent.mkdir(parents=True, exist_ok=True)
    if manifest_path is not None:
        manifest_path.parent.mkdir(parents=True, exist_ok=True)

    temporary = None
    report_temporary = None
    manifest_temporary = None
    recalculated = False
    neware_source = None
    manifest = None
    updated_manifest = None
    neware_assignments = ()
    neware_new_ids = ()
    neware_retained_ids = ()
    neware_absent_labels = ()
    try:
        if neware_source_path is not None:
            try:
                neware_source = read_neware_source(neware_source_path)
                manifest = load_neware_manifest(manifest_path)
            except NewareInputError as exc:
                raise CandidateBuildError(str(exc)) from exc
        elif manifest_path is not None:
            try:
                manifest = load_neware_manifest(manifest_path)
            except NewareInputError as exc:
                raise CandidateBuildError(str(exc)) from exc

        with tempfile.NamedTemporaryFile(
            mode="wb",
            suffix=".xlsx",
            prefix=f".{candidate.stem}.",
            dir=candidate.parent,
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
        shutil.copy2(database, temporary)

        workbook = _load_workbook_for_edit(temporary)
        try:
            if MIRROR_SHEET_NAME not in workbook.sheetnames or DATABASE_SHEET_NAME not in workbook.sheetnames:
                raise CandidateBuildError("database must contain Slurry and db_table")
            slurry = workbook[MIRROR_SHEET_NAME]
            database_sheet = workbook[DATABASE_SHEET_NAME]

            source_rows, filtered = _read_filtered_source(
                source,
                max_columns=slurry.max_column,
            )
            existing_ids, existing_duplicate_ids = _existing_slurry_ids(slurry)
            manifest_ids = (
                tuple(entry.record_id for entry in manifest.entries.values())
                if manifest is not None
                else ()
            )
            existing_database_ids = _database_ids_against_slurry(
                database_sheet,
                slurry,
                allowed_neware_ids=manifest_ids,
            )
            if manifest is not None:
                _validate_existing_neware_rows(database_sheet, manifest)
            existing_set = set(existing_ids)
            filtered_ids = tuple(record_id for record_id, _ in filtered)
            filtered_set = set(filtered_ids)
            retained_ids = tuple(record_id for record_id in filtered_ids if record_id in existing_set)
            new_rows = tuple((record_id, row) for record_id, row in filtered if record_id not in existing_set)
            new_ids = tuple(record_id for record_id, _ in new_rows)
            absent_existing_ids = tuple(sorted(existing_set - filtered_set))

            if neware_source is not None:
                try:
                    (
                        updated_manifest,
                        neware_assignments,
                        neware_new_ids,
                        neware_retained_ids,
                        neware_absent_labels,
                    ) = _prepare_neware_records(
                        neware_source,
                        manifest,
                        existing_database_ids=existing_database_ids,
                        existing_slurry_ids=existing_ids,
                        legacy_source_ids=filtered_ids,
                    )
                except NewareInputError as exc:
                    raise CandidateBuildError(str(exc)) from exc
            else:
                updated_manifest = manifest

            initial_database_rows = database_sheet.max_row
            initial_database_columns = database_sheet.max_column
            existing_database_values = _worksheet_values(
                database_sheet,
                max_row=initial_database_rows,
                max_column=initial_database_columns,
            )
            manual_values = tuple(
                tuple(database_sheet.cell(row_number, column).value for column in MANUAL_COLUMNS)
                for row_number in range(DATABASE_DATA_START_ROW, initial_database_rows + 1)
            )
            initial_slurry_rows = slurry.max_row
            initial_slurry_columns = slurry.max_column
            existing_slurry_values = _worksheet_values(
                slurry,
                max_row=initial_slurry_rows,
                max_column=initial_slurry_columns,
            )

            appended_slurry_rows = []
            appended_database_rows = []
            neware_appended_rows = []
            for record_id, row in new_rows:
                row_number = max(MIRROR_DATA_START_ROW - 1, slurry.max_row) + 1
                appended_slurry_rows.append(row_number)
                for column, value in enumerate(row, start=1):
                    slurry.cell(row_number, column, value)

                database_row = max(DATABASE_DATA_START_ROW - 1, database_sheet.max_row) + 1
                appended_database_rows.append(database_row)
                if cellpy_ready:
                    for column, source_column in DB_TABLE_FORMULA_COLUMNS.items():
                        database_sheet.cell(
                            database_row,
                            column,
                            f"=Slurry!{source_column}{row_number}",
                        )
                    database_sheet.cell(database_row, 27, f"=Z{database_row}")
                else:
                    database_sheet.cell(database_row, 1, record_id)
                for column, system_name in SYSTEM_COLUMNS.items():
                    database_sheet.cell(database_row, column, SYSTEM_VALUES[system_name])

            for neware_row, record_id, should_append in neware_assignments:
                if not should_append:
                    continue
                database_row = max(DATABASE_DATA_START_ROW - 1, database_sheet.max_row) + 1
                database_sheet.cell(database_row, 1, record_id)
                mapped_values = neware_row.mapped_values()
                for column, field in NEWARE_DB_COLUMN_MAP.items():
                    database_sheet.cell(database_row, column, mapped_values[field])
                for column, system_name in SYSTEM_COLUMNS.items():
                    database_sheet.cell(database_row, column, SYSTEM_VALUES[system_name])
                neware_appended_rows.append((database_row, record_id, neware_row))

            if new_rows or neware_appended_rows:
                formula_snapshot = _formula_cache_snapshot(temporary)
                _save_workbook_for_edit(workbook, temporary)
                _restore_formula_caches(temporary, formula_snapshot)
        finally:
            workbook.close()

        _verify_candidate(
            temporary,
            initial_slurry_rows=initial_slurry_rows,
            initial_slurry_columns=initial_slurry_columns,
            existing_slurry_values=existing_slurry_values,
            initial_database_rows=initial_database_rows,
            initial_database_columns=initial_database_columns,
            existing_database_values=existing_database_values,
            manual_values=manual_values,
            appended_slurry_rows=tuple(appended_slurry_rows),
            appended_database_rows=tuple(appended_database_rows),
            new_ids=new_ids,
            neware_appended_rows=tuple(neware_appended_rows),
            cellpy_ready=cellpy_ready,
        )

        if cellpy_ready:
            _recalculate_with_excel(temporary)
            _verify_cellpy_ready(
                temporary,
                appended_database_rows=tuple(appended_database_rows),
                new_ids=new_ids,
            )
            recalculated = True

        result = CandidateReport(
            source_path=str(source),
            database_path=str(database),
            candidate_path=str(candidate),
            source_rows=source_rows,
            filtered_rows=len(filtered),
            existing_slurry_rows=len(existing_ids),
            retained_ids=retained_ids,
            new_ids=new_ids,
            absent_existing_ids=absent_existing_ids,
            existing_duplicate_ids=existing_duplicate_ids,
            cellpy_ready=cellpy_ready,
            recalculated=recalculated,
            neware_source_path=(str(neware_source.source_path) if neware_source is not None else None),
            neware_manifest_path=(str(manifest_path) if neware_source is not None else None),
            neware_source_rows=(neware_source.source_rows if neware_source is not None else 0),
            neware_usable_rows=(len(neware_source.rows) if neware_source is not None else 0),
            neware_placeholder_rows=(neware_source.placeholder_rows if neware_source is not None else ()),
            neware_retained_ids=tuple(neware_retained_ids),
            neware_new_ids=tuple(neware_new_ids),
            neware_absent_labels=tuple(neware_absent_labels),
        )
        if report is not None:
            report_temporary = _stage_report(report, result)
        if (
            neware_source is not None
            and updated_manifest is not None
            and manifest is not None
            and updated_manifest.document() != manifest.document()
        ):
            manifest_temporary = _stage_manifest(manifest_path, updated_manifest)
        _publish_outputs(
            temporary,
            candidate,
            report_temporary,
            report,
            manifest_temporary,
            manifest_path,
        )
        temporary = None
        report_temporary = None
        manifest_temporary = None
        return result
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)
        if report_temporary is not None:
            report_temporary.unlink(missing_ok=True)
        if manifest_temporary is not None:
            manifest_temporary.unlink(missing_ok=True)
