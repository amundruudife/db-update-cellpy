"""Typed, candidate-only input handling for a Neware ``test_log`` workbook.

The Neware workbook is deliberately treated as a small independent source.  Only
the named ``test_log`` sheet and the fields in :data:`NEWARE_PAYLOAD_FIELDS` are
part of this contract; all other sheets and columns are ignored.
"""

from __future__ import annotations

import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook


NEWARE_SHEET_NAME = "test_log"
NEWARE_HEADER_ROW = 1
NEWARE_TYPE_ROW = 2
NEWARE_DATA_START_ROW = 3

NEWARE_REQUIRED_FIELDS = (
    "Test No",
    "Cell ID",
    "Cell batch",
    "Project",
    "Cell type",
    "Cell test label",
    "Cell capacity (Ah)",
    "Test Schedule",
    "Test temp (°C)",
    "Comment",
)
NEWARE_PAYLOAD_FIELDS = NEWARE_REQUIRED_FIELDS

# These are the two template rows in the supplied Neware export.  The leading
# space is intentional and is retained from the source value.  The unpadded
# spelling is accepted as the same explicit template marker for sanitized
# copies of that workbook.
NEWARE_PLACEHOLDER_CELL_IDS = frozenset(
    {
        " Cell ID or Cell serial number",
        "Cell ID or Cell serial number",
    }
)


class NewareInputError(ValueError):
    """Raised when a Neware workbook or ID manifest is unsafe to use."""


def _blank_as_none(value):
    return None if value == "" else value


def _validate_json_value(value, *, context: str):
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            raise NewareInputError(f"invalid non-finite {context}: {value!r}")
        return value
    raise NewareInputError(f"unsupported {context} type: {type(value).__name__}")


@dataclass(frozen=True)
class NewareRow:
    """One usable Neware row, with source typing preserved."""

    row_number: int
    test_no: object
    cell_id: object
    cell_batch: object
    project: object
    cell_type: object
    cell_test_label: str
    cell_capacity_ah: object
    test_schedule: object
    test_temp_c: object
    comment: object

    def payload(self) -> dict[str, object]:
        """Return the complete identity/payload contract for manifest checks."""

        return {
            "Test No": self.test_no,
            "Cell ID": self.cell_id,
            "Cell batch": self.cell_batch,
            "Project": self.project,
            "Cell type": self.cell_type,
            "Cell test label": self.cell_test_label,
            "Cell capacity (Ah)": self.cell_capacity_ah,
            "Test Schedule": self.test_schedule,
            "Test temp (°C)": self.test_temp_c,
            "Comment": self.comment,
        }

    def mapped_values(self) -> dict[str, object]:
        """Return only the approved target-field mappings."""

        return {
            "label": self.cell_id,
            "batch": self.cell_batch,
            "project": self.project,
            "cell_type": self.cell_type,
            "cell": self.cell_test_label,
            "nominal_capacity": self.cell_capacity_ah,
            "schedule": self.test_schedule,
            "temperature": self.test_temp_c,
            "comment_general": self.comment,
        }


@dataclass(frozen=True)
class NewareSource:
    """Parsed Neware rows and the explicitly excluded template rows."""

    source_path: Path
    source_rows: int
    rows: tuple[NewareRow, ...]
    placeholder_rows: tuple[int, ...]


def _header_indexes(sheet) -> dict[str, int]:
    headers = tuple(sheet.cell(NEWARE_HEADER_ROW, column).value for column in range(1, sheet.max_column + 1))
    indexes: dict[str, int] = {}
    for required in NEWARE_REQUIRED_FIELDS:
        matches = [index for index, header in enumerate(headers) if header == required]
        if len(matches) != 1:
            raise NewareInputError(
                f"test_log must contain exactly one {required!r} header"
            )
        indexes[required] = matches[0]
    return indexes


def read_neware_source(source_path: str | Path) -> NewareSource:
    """Read only ``test_log`` and return validated usable rows.

    Row 2 is the Neware type-description row, not a record.  Empty records are
    ignored.  The two explicit template markers are counted and reported rather
    than being treated as records.
    """

    path = Path(source_path).resolve(strict=False)
    if not path.is_file():
        raise FileNotFoundError(f"Neware source workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if NEWARE_SHEET_NAME not in workbook.sheetnames:
            raise NewareInputError("Neware source workbook must contain the 'test_log' sheet")
        sheet = workbook[NEWARE_SHEET_NAME]
        indexes = _header_indexes(sheet)
        rows: list[NewareRow] = []
        placeholder_rows: list[int] = []
        source_rows = 0
        for row_number, raw_row in enumerate(
            sheet.iter_rows(min_row=NEWARE_DATA_START_ROW, values_only=True),
            start=NEWARE_DATA_START_ROW,
        ):
            values = {
                field: _blank_as_none(raw_row[index]) if index < len(raw_row) else None
                for field, index in indexes.items()
            }
            if not any(value is not None for value in values.values()):
                continue
            source_rows += 1
            cell_id = values["Cell ID"]
            if cell_id in NEWARE_PLACEHOLDER_CELL_IDS:
                placeholder_rows.append(row_number)
                continue

            label = values["Cell test label"]
            if not isinstance(label, str) or not label.strip():
                raise NewareInputError(
                    f"usable test_log row {row_number} has a blank Cell test label"
                )
            typed_values = {
                field: _validate_json_value(value, context=f"{field} at row {row_number}")
                for field, value in values.items()
            }
            rows.append(
                NewareRow(
                    row_number=row_number,
                    test_no=typed_values["Test No"],
                    cell_id=typed_values["Cell ID"],
                    cell_batch=typed_values["Cell batch"],
                    project=typed_values["Project"],
                    cell_type=typed_values["Cell type"],
                    cell_test_label=label,
                    cell_capacity_ah=typed_values["Cell capacity (Ah)"],
                    test_schedule=typed_values["Test Schedule"],
                    test_temp_c=typed_values["Test temp (°C)"],
                    comment=typed_values["Comment"],
                )
            )

        labels = [row.cell_test_label for row in rows]
        duplicates = sorted({label for label in labels if labels.count(label) > 1})
        if duplicates:
            raise NewareInputError(
                "duplicate Cell test label: " + ", ".join(repr(label) for label in duplicates)
            )
        return NewareSource(
            source_path=path,
            source_rows=source_rows,
            rows=tuple(rows),
            placeholder_rows=tuple(placeholder_rows),
        )
    finally:
        workbook.close()


@dataclass(frozen=True)
class NewareManifestEntry:
    record_id: int
    payload: dict[str, object]


@dataclass(frozen=True)
class NewareManifest:
    """Validated natural-key-to-database-ID assignments."""

    path: Path
    entries: dict[str, NewareManifestEntry]

    def document(self) -> dict[str, object]:
        return {
            "version": 1,
            "natural_key": "Cell test label",
            "entries": {
                label: {
                    "id": entry.record_id,
                    "payload": entry.payload,
                }
                for label, entry in sorted(self.entries.items())
            },
        }


def _positive_manifest_id(value, *, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise NewareInputError(f"invalid manifest ID for {label!r}: {value!r}")
    return value


def load_neware_manifest(path: str | Path) -> NewareManifest:
    """Load and strictly validate a Neware ID manifest, or return an empty one."""

    manifest_path = Path(path).resolve(strict=False)
    if not manifest_path.exists():
        return NewareManifest(path=manifest_path, entries={})
    try:
        document = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise NewareInputError(f"could not read Neware ID manifest: {manifest_path}") from exc
    if not isinstance(document, dict) or document.get("version") != 1:
        raise NewareInputError("Neware ID manifest has an unsupported version")
    if document.get("natural_key") != "Cell test label":
        raise NewareInputError("Neware ID manifest has an invalid natural key")
    raw_entries = document.get("entries")
    if not isinstance(raw_entries, dict):
        raise NewareInputError("Neware ID manifest entries must be an object")

    entries: dict[str, NewareManifestEntry] = {}
    ids: dict[int, str] = {}
    for label, raw_entry in raw_entries.items():
        if not isinstance(label, str) or not label.strip():
            raise NewareInputError("Neware ID manifest contains a blank natural key")
        if not isinstance(raw_entry, dict):
            raise NewareInputError(f"manifest entry for {label!r} must be an object")
        record_id = _positive_manifest_id(raw_entry.get("id"), label=label)
        other_label = ids.get(record_id)
        if other_label is not None:
            raise NewareInputError(
                f"Neware ID manifest ID collision: {record_id} maps to {other_label!r} and {label!r}"
            )
        payload = raw_entry.get("payload")
        if not isinstance(payload, dict) or set(payload) != set(NEWARE_PAYLOAD_FIELDS):
            raise NewareInputError(f"manifest payload for {label!r} has the wrong fields")
        if payload.get("Cell test label") != label:
            raise NewareInputError(f"manifest payload label mismatch for {label!r}")
        for field in NEWARE_PAYLOAD_FIELDS:
            _validate_json_value(payload[field], context=f"manifest {field} for {label!r}")
        entries[label] = NewareManifestEntry(record_id=record_id, payload=dict(payload))
        ids[record_id] = label
    return NewareManifest(path=manifest_path, entries=entries)


def manifest_with_entries(
    manifest: NewareManifest,
    entries: Mapping[str, NewareManifestEntry],
) -> NewareManifest:
    """Return a validated immutable manifest value with replacement entries."""

    ids: dict[int, str] = {}
    copied: dict[str, NewareManifestEntry] = {}
    for label, entry in entries.items():
        if label in copied:
            raise NewareInputError(f"duplicate Neware manifest label: {label!r}")
        if entry.record_id in ids and ids[entry.record_id] != label:
            raise NewareInputError(f"Neware ID collision: {entry.record_id}")
        if entry.payload.get("Cell test label") != label:
            raise NewareInputError(f"manifest payload label mismatch for {label!r}")
        ids[entry.record_id] = label
        copied[label] = NewareManifestEntry(entry.record_id, dict(entry.payload))
    return NewareManifest(path=manifest.path, entries=copied)


__all__ = [
    "NEWARE_DATA_START_ROW",
    "NEWARE_PAYLOAD_FIELDS",
    "NEWARE_PLACEHOLDER_CELL_IDS",
    "NEWARE_REQUIRED_FIELDS",
    "NEWARE_SHEET_NAME",
    "NewareInputError",
    "NewareManifest",
    "NewareManifestEntry",
    "NewareRow",
    "NewareSource",
    "load_neware_manifest",
    "manifest_with_entries",
    "read_neware_source",
]
